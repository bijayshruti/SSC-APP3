"""
STAFF SELECTION COMMISSION (ER), KOLKATA
Centre Coordinator & Flying Squad Allocation System
Streamlit Web Application
Designed by Bijay Paswan
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
import sys
import logging
import warnings
import hashlib
import base64
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ============================================================================
# CONSTANTS AND CONFIGURATION
# ============================================================================

# Create necessary directories
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
BACKUP_DIR = DATA_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

# File paths
CONFIG_FILE = DATA_DIR / "config.json"
DATA_FILE = DATA_DIR / "allocations_data.json"
REFERENCE_FILE = DATA_DIR / "allocation_references.json"
DELETED_RECORDS_FILE = DATA_DIR / "deleted_records.json"
LOGFILE = DATA_DIR / "app.log"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename=LOGFILE
)

# Default remuneration rates
DEFAULT_RATES = {
    'multiple_shifts': 750,
    'single_shift': 450,
    'mock_test': 450,
    'ey_personnel': 5000
}

# Default colors for UI
COLORS = {
    'primary': '#4169e1',
    'danger': '#dc143c',
    'success': '#3cb371',
    'warning': '#ff8c00',
    'info': '#20b2aa',
    'dark': '#2c3e50',
    'light': '#f8f9fa'
}

# ============================================================================
# SESSION STATE MANAGEMENT - COMPLETE INITIALIZATION
# ============================================================================

def initialize_session_state():
    """Initialize all session state variables"""
    # Data storage
    default_values = {
        'io_df': pd.DataFrame(),
        'venue_df': pd.DataFrame(),
        'ey_df': pd.DataFrame(),
        'allocation': [],
        'ey_allocation': [],
        'deleted_records': [],
        
        # Configuration
        'remuneration_rates': DEFAULT_RATES.copy(),
        'exam_data': {},
        'allocation_references': {},
        
        # Current state
        'current_exam_key': "",
        'exam_name': "",
        'exam_year': "",
        'mock_test_mode': False,
        'ey_allocation_mode': False,
        
        # UI state
        'selected_venue': "",
        'selected_role': "Centre Coordinator",
        'selected_dates': [],
        'selected_io': "",
        'selected_ey': "",
        
        # Date selection state
        'date_selection_state': {},
        'expanded_dates': {},
        
        # EY date selection state
        'ey_date_selection_state': {},
        'ey_expanded_dates': {},
        
        # File upload tracking
        'io_master_loaded': False,
        'venue_master_loaded': False,
        'ey_master_loaded': False,
        
        # Current allocation state (for reference handling)
        'current_allocation_person': None,
        'current_allocation_area': None,
        'current_allocation_role': None,
        'current_allocation_type': None,
        'current_allocation_ey_row': None,
        
        # UI flags
        'show_io_upload': False,
        'show_venue_upload': False,
        'show_ey_upload': False,
        'creating_new_ref_IO': False,
        'creating_new_ref_EY Personnel': False,
        
        # Deletion management
        'selected_allocation_for_deletion': [],
        'selected_ey_allocation_for_deletion': [],
        'show_delete_confirmation': False,
        
        # Current menu
        'menu': 'dashboard'
    }
    
    # Initialize all values
    for key, default_value in default_values.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

# ============================================================================
# DATA PERSISTENCE FUNCTIONS - FIXED
# ============================================================================

def load_all_data():
    """Load all data from files"""
    try:
        # Load config
        if CONFIG_FILE.exists():
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                st.session_state.remuneration_rates = config.get('remuneration_rates', DEFAULT_RATES.copy())
                logging.info(f"Config loaded: {len(config)} items")
        
        # Load exam data - handle both old and new formats
        if DATA_FILE.exists():
            with open(DATA_FILE, 'r') as f:
                exam_data = json.load(f)
                # Convert loaded data to proper format
                st.session_state.exam_data = {}
                for exam_key, data in exam_data.items():
                    if isinstance(data, dict):
                        # New format: {'io_allocations': [], 'ey_allocations': []}
                        st.session_state.exam_data[exam_key] = data
                    else:
                        # Old format: direct list of allocations
                        st.session_state.exam_data[exam_key] = {
                            'io_allocations': data,
                            'ey_allocations': []
                        }
                logging.info(f"Exam data loaded: {len(exam_data)} exams")
        
        # Load reference data
        if REFERENCE_FILE.exists():
            with open(REFERENCE_FILE, 'r') as f:
                ref_data = json.load(f)
                st.session_state.allocation_references = ref_data
                logging.info(f"References loaded: {len(ref_data)} items")
        
        # Load deleted records
        if DELETED_RECORDS_FILE.exists():
            with open(DELETED_RECORDS_FILE, 'r') as f:
                deleted_data = json.load(f)
                st.session_state.deleted_records = deleted_data
                logging.info(f"Deleted records loaded: {len(deleted_data)} items")
        
        # Load current allocations from current exam
        if st.session_state.current_exam_key and st.session_state.current_exam_key in st.session_state.exam_data:
            exam_data = st.session_state.exam_data[st.session_state.current_exam_key]
            st.session_state.allocation = exam_data.get('io_allocations', [])
            st.session_state.ey_allocation = exam_data.get('ey_allocations', [])
            logging.info(f"Current allocations loaded: {len(st.session_state.allocation)} IO, {len(st.session_state.ey_allocation)} EY")
        
        return True
    except Exception as e:
        logging.error(f"Error loading data: {str(e)}")
        st.error(f"Error loading data: {str(e)}")
        return False

def save_all_data():
    """Save all data to files"""
    try:
        # Update current exam data before saving
        if st.session_state.current_exam_key:
            if st.session_state.current_exam_key not in st.session_state.exam_data:
                st.session_state.exam_data[st.session_state.current_exam_key] = {
                    'io_allocations': [],
                    'ey_allocations': []
                }
            
            # Update allocations for current exam
            st.session_state.exam_data[st.session_state.current_exam_key]['io_allocations'] = st.session_state.allocation
            st.session_state.exam_data[st.session_state.current_exam_key]['ey_allocations'] = st.session_state.ey_allocation
        
        # Save config
        config = {
            'remuneration_rates': st.session_state.remuneration_rates,
            'last_saved': datetime.now().isoformat()
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4, default=str)
        
        # Save exam data
        with open(DATA_FILE, 'w') as f:
            json.dump(st.session_state.exam_data, f, indent=4, default=str)
        
        # Save reference data
        with open(REFERENCE_FILE, 'w') as f:
            json.dump(st.session_state.allocation_references, f, indent=4, default=str)
        
        # Save deleted records
        with open(DELETED_RECORDS_FILE, 'w') as f:
            json.dump(st.session_state.deleted_records, f, indent=4, default=str)
        
        logging.info("All data saved successfully")
        return True
    except Exception as e:
        logging.error(f"Error saving data: {str(e)}")
        st.error(f"Error saving data: {str(e)}")
        return False

# ============================================================================
# BACKUP MANAGEMENT
# ============================================================================

def create_backup(description=""):
    """Create a backup of current data"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}"
        if description:
            backup_name += f"_{description.replace(' ', '_')}"
        backup_file = BACKUP_DIR / f"{backup_name}.json"
        
        # Get current data
        backup_data = {
            'timestamp': datetime.now().isoformat(),
            'description': description,
            'exam_data': st.session_state.exam_data,
            'allocation_references': st.session_state.allocation_references,
            'remuneration_rates': st.session_state.remuneration_rates,
            'deleted_records': st.session_state.deleted_records,
            'current_exam_key': st.session_state.current_exam_key
        }
        
        with open(backup_file, 'w') as f:
            json.dump(backup_data, f, indent=4, default=str)
        
        logging.info(f"Backup created: {backup_file.name}")
        return backup_file
    except Exception as e:
        logging.error(f"Error creating backup: {str(e)}")
        return None

def restore_from_backup(backup_file):
    """Restore data from backup file"""
    try:
        with open(backup_file, 'r') as f:
            backup_data = json.load(f)
        
        # Restore data
        st.session_state.exam_data = backup_data.get('exam_data', {})
        st.session_state.allocation_references = backup_data.get('allocation_references', {})
        st.session_state.remuneration_rates = backup_data.get('remuneration_rates', DEFAULT_RATES.copy())
        st.session_state.deleted_records = backup_data.get('deleted_records', [])
        st.session_state.current_exam_key = backup_data.get('current_exam_key', "")
        
        # Clear current allocations
        st.session_state.allocation = []
        st.session_state.ey_allocation = []
        
        # Load allocations if exam key exists
        if st.session_state.current_exam_key and st.session_state.current_exam_key in st.session_state.exam_data:
            exam_data = st.session_state.exam_data[st.session_state.current_exam_key]
            st.session_state.allocation = exam_data.get('io_allocations', [])
            st.session_state.ey_allocation = exam_data.get('ey_allocations', [])
        
        # Save restored data
        save_all_data()
        
        logging.info(f"Data restored from backup: {backup_file.name}")
        return True
    except Exception as e:
        logging.error(f"Error restoring from backup: {str(e)}")
        return False

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def load_default_master_data():
    """Load default master data for demonstration"""
    # Default IO data
    default_io_data = {
        'NAME': [
            'John Doe', 'Jane Smith', 'Robert Johnson', 
            'Emily Davis', 'Michael Wilson', 'Sarah Brown',
            'David Miller', 'Lisa Anderson', 'James Taylor',
            'Maria Thomas'
        ],
        'AREA': [
            'Kolkata', 'Howrah', 'Hooghly', 'Nadia', 
            'North 24 Parganas', 'South 24 Parganas',
            'Bardhaman', 'Birbhum', 'Bankura', 'Purulia'
        ],
        'CENTRE_CODE': ['1001', '1002', '1003', '2001', '2002', '3001', '3002', '4001', '4002', '5001'],
        'MOBILE': [
            '9876543210', '9876543211', '9876543212', '9876543213', '9876543214',
            '9876543215', '9876543216', '9876543217', '9876543218', '9876543219'
        ],
        'EMAIL': [
            'john@example.com', 'jane@example.com', 'robert@example.com',
            'emily@example.com', 'michael@example.com', 'sarah@example.com',
            'david@example.com', 'lisa@example.com', 'james@example.com',
            'maria@example.com'
        ],
        'DESIGNATION': [
            'Assistant Commissioner', 'Deputy Commissioner', 'Assistant Commissioner',
            'Section Officer', 'Assistant Section Officer', 'Section Officer',
            'Deputy Commissioner', 'Assistant Commissioner', 'Section Officer',
            'Assistant Section Officer'
        ],
        'BANK_NAME': ['SBI', 'PNB', 'BOB', 'SBI', 'PNB', 'BOB', 'SBI', 'PNB', 'BOB', 'SBI'],
        'ACCOUNT_NUMBER': ['1234567890', '2345678901', '3456789012', '4567890123', '5678901234',
                          '6789012345', '7890123456', '8901234567', '9012345678', '0123456789'],
        'IFSC_CODE': ['SBIN0001234', 'PNBN0012345', 'BARB0XXXXXX', 'SBIN0001234', 'PNBN0012345',
                     'BARB0XXXXXX', 'SBIN0001234', 'PNBN0012345', 'BARB0XXXXXX', 'SBIN0001234']
    }
    
    st.session_state.io_df = pd.DataFrame(default_io_data)
    st.session_state.io_master_loaded = True
    
    # Default venue data
    venues = [
        'Kolkata Main Centre', 'Howrah Centre', 'Hooghly Centre',
        'Nadia Centre', 'North 24 Parganas Centre', 'South 24 Parganas Centre',
        'Bardhaman Centre', 'Birbhum Centre', 'Bankura Centre', 'Purulia Centre'
    ]
    
    venue_data = []
    start_date = datetime.now().date()
    
    for venue_idx, venue_name in enumerate(venues):
        centre_code = f'100{venue_idx + 1}'
        for day_offset in range(5):  # 5 days
            date = start_date + timedelta(days=day_offset)
            date_str = date.strftime("%d-%m-%Y")
            for shift in ['Morning', 'Afternoon', 'Evening']:
                venue_data.append({
                    'VENUE': str(venue_name),
                    'DATE': str(date_str),
                    'SHIFT': str(shift),
                    'CENTRE_CODE': str(centre_code),
                    'ADDRESS': f'Address for {venue_name}',
                    'CENTRE NAME': str(venue_name),
                    'STATE': 'West Bengal',
                    'DISTRICT': str(venue_name.split()[0]),
                    'CAPACITY': 500 + (venue_idx * 100)
                })
    
    st.session_state.venue_df = pd.DataFrame(venue_data)
    st.session_state.venue_master_loaded = True
    
    # Default EY data
    default_ey_data = {
        'NAME': [
            'Dr. Amit Sharma', 'Prof. Priya Gupta', 'Ms. Anjali Chatterjee',
            'Mr. Rajesh Banerjee', 'Dr. Sunita Das', 'Prof. Ravi Kumar',
            'Ms. Meera Sen', 'Mr. Arjun Roy', 'Dr. Neha Verma',
            'Prof. Sanjay Mishra'
        ],
        'MOBILE': [
            '9876543201', '9876543202', '9876543203', '9876543204', '9876543205',
            '9876543206', '9876543207', '9876543208', '9876543209', '9876543210'
        ],
        'EMAIL': [
            'sharma@example.com', 'gupta@example.com', 'chatterjee@example.com',
            'banerjee@example.com', 'das@example.com', 'kumar@example.com',
            'sen@example.com', 'roy@example.com', 'verma@example.com', 'mishra@example.com'
        ],
        'ID_NUMBER': [f'EY00{i+1}' for i in range(10)],
        'DESIGNATION': [
            'Professor', 'Associate Professor', 'Assistant Professor',
            'Lecturer', 'Professor', 'Associate Professor',
            'Assistant Professor', 'Lecturer', 'Professor', 'Associate Professor'
        ],
        'DEPARTMENT': [
            'Mathematics', 'Physics', 'Chemistry', 'English',
            'History', 'Computer Science', 'Economics', 'Political Science',
            'Biology', 'Statistics'
        ],
        'UNIVERSITY': [
            'University of Calcutta', 'Jadavpur University', 'Presidency University',
            'University of Calcutta', 'Jadavpur University', 'Presidency University',
            'University of Calcutta', 'Jadavpur University', 'Presidency University',
            'University of Calcutta'
        ]
    }
    
    st.session_state.ey_df = pd.DataFrame(default_ey_data)
    st.session_state.ey_master_loaded = True
    
    st.success("Default master data loaded successfully!")

def get_file_download_link(df, filename, filetype='csv'):
    """Generate a download link for a dataframe"""
    if filetype == 'csv':
        data = df.to_csv(index=False)
        mime_type = 'text/csv'
        extension = 'csv'
    elif filetype == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
        data = output.getvalue()
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        extension = 'xlsx'
    else:
        return None
    
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:{mime_type};base64,{b64}" download="{filename}.{extension}">Download {filename}</a>'
    return href

def check_allocation_conflict(person_name, date, shift, venue, role, allocation_type):
    """Check for allocation conflicts"""
    if allocation_type == "IO":
        # Check for exact duplicate
        duplicate = any(
            alloc['IO Name'] == person_name and 
            alloc['Date'] == date and 
            alloc['Shift'] == shift and 
            alloc['Venue'] == venue and 
            alloc['Role'] == role
            for alloc in st.session_state.allocation
        )
        
        if duplicate:
            return True, f"Duplicate allocation found! {person_name} is already allocated to {venue} on {date} ({shift}) as {role}."
        
        # For Centre Coordinator: Cannot be at multiple venues same date/shift
        if role == "Centre Coordinator":
            conflict = any(
                alloc['IO Name'] == person_name and 
                alloc['Date'] == date and 
                alloc['Shift'] == shift and 
                alloc['Venue'] != venue and
                alloc['Role'] == "Centre Coordinator"
                for alloc in st.session_state.allocation
            )
            if conflict:
                existing_venue = next(
                    alloc['Venue'] for alloc in st.session_state.allocation 
                    if alloc['IO Name'] == person_name and 
                       alloc['Date'] == date and 
                       alloc['Shift'] == shift and
                       alloc['Role'] == "Centre Coordinator"
                )
                return True, f"Centre Coordinator conflict! {person_name} is already allocated to {existing_venue} on {date} ({shift})."
        
        # For Flying Squad: Allow multiple venues but warn
        elif role == "Flying Squad":
            existing_venues = [
                alloc['Venue'] for alloc in st.session_state.allocation 
                if alloc['IO Name'] == person_name and 
                   alloc['Date'] == date and 
                   alloc['Shift'] == shift and
                   alloc['Role'] == "Flying Squad"
            ]
            if existing_venues:
                return False, f"Warning: {person_name} is already allocated to {', '.join(existing_venues)} on {date} ({shift}). Do you want to assign to additional venue {venue}?"
    
    elif allocation_type == "EY":
        # Check for exact duplicate
        duplicate = any(
            alloc['EY Personnel'] == person_name and 
            alloc['Date'] == date and 
            alloc['Shift'] == shift and 
            alloc['Venue'] == venue
            for alloc in st.session_state.ey_allocation
        )
        
        if duplicate:
            return True, f"Duplicate EY allocation found! {person_name} is already allocated to {venue} on {date} ({shift})."
        
        # EY Personnel: Cannot be at multiple venues same date/shift
        conflict = any(
            alloc['EY Personnel'] == person_name and 
            alloc['Date'] == date and 
            alloc['Shift'] == shift and 
            alloc['Venue'] != venue
            for alloc in st.session_state.ey_allocation
        )
        
        if conflict:
            existing_venue = next(
                alloc['Venue'] for alloc in st.session_state.ey_allocation 
                if alloc['EY Personnel'] == person_name and 
                   alloc['Date'] == date and 
                   alloc['Shift'] == shift
            )
            return True, f"EY Personnel conflict! {person_name} is already allocated to {existing_venue} on {date} ({shift})."
    
    return False, ""

def delete_allocation(allocation_index, allocation_type="IO"):
    """Delete an allocation and move to deleted records"""
    try:
        if allocation_type == "IO":
            if allocation_index < 0 or allocation_index >= len(st.session_state.allocation):
                return False
            
            # Get allocation record
            allocation = st.session_state.allocation[allocation_index]
            
            # Add to deleted records
            deleted_record = {
                **allocation,
                'Deletion Timestamp': datetime.now().isoformat(),
                'Deletion Reason': 'Manual deletion',
                'Type': 'IO'
            }
            st.session_state.deleted_records.append(deleted_record)
            
            # Remove from allocations
            del st.session_state.allocation[allocation_index]
            
            # Update serial numbers
            for idx, alloc in enumerate(st.session_state.allocation):
                alloc['Sl. No.'] = idx + 1
            
            # Update exam data
            if st.session_state.current_exam_key in st.session_state.exam_data:
                st.session_state.exam_data[st.session_state.current_exam_key]['io_allocations'] = st.session_state.allocation
            
            logging.info(f"Deleted IO allocation: {allocation.get('IO Name', 'Unknown')}")
            
        elif allocation_type == "EY":
            if allocation_index < 0 or allocation_index >= len(st.session_state.ey_allocation):
                return False
            
            # Get allocation record
            allocation = st.session_state.ey_allocation[allocation_index]
            
            # Add to deleted records
            deleted_record = {
                **allocation,
                'Deletion Timestamp': datetime.now().isoformat(),
                'Deletion Reason': 'Manual deletion',
                'Type': 'EY Personnel'
            }
            st.session_state.deleted_records.append(deleted_record)
            
            # Remove from allocations
            del st.session_state.ey_allocation[allocation_index]
            
            # Update serial numbers
            for idx, alloc in enumerate(st.session_state.ey_allocation):
                alloc['Sl. No.'] = idx + 1
            
            # Update exam data
            if st.session_state.current_exam_key in st.session_state.exam_data:
                st.session_state.exam_data[st.session_state.current_exam_key]['ey_allocations'] = st.session_state.ey_allocation
            
            logging.info(f"Deleted EY allocation: {allocation.get('EY Personnel', 'Unknown')}")
        
        save_all_data()
        return True
        
    except Exception as e:
        logging.error(f"Error deleting allocation: {str(e)}")
        return False

def bulk_delete_allocations(indices, allocation_type="IO"):
    """Delete multiple allocations in bulk"""
    try:
        # Sort indices in descending order to avoid index shifting issues
        indices.sort(reverse=True)
        
        deleted_count = 0
        for idx in indices:
            if delete_allocation(idx, allocation_type):
                deleted_count += 1
        
        return deleted_count
    except Exception as e:
        logging.error(f"Error in bulk deletion: {str(e)}")
        return 0

# ============================================================================
# REFERENCE MANAGEMENT - FIXED
# ============================================================================

def get_or_create_reference(allocation_type):
    """Get existing reference or create new one"""
    exam_key = st.session_state.current_exam_key
    if not exam_key:
        st.warning("Please select or create an exam first")
        return None
    
    if exam_key not in st.session_state.allocation_references:
        st.session_state.allocation_references[exam_key] = {}
    
    role_key = allocation_type
    
    # Check if reference exists
    if role_key in st.session_state.allocation_references[exam_key]:
        existing_ref = st.session_state.allocation_references[exam_key][role_key]
        
        # Create columns for choice
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button(f"📝 Use Existing Reference", key=f"use_existing_{allocation_type}"):
                return existing_ref
        
        with col2:
            if st.button(f"🆕 Create New Reference", key=f"new_ref_{allocation_type}"):
                st.session_state[f"creating_new_ref_{allocation_type}"] = True
                st.rerun()
        
        # Display existing reference info
        st.info(f"**Existing Reference:** Order No. {existing_ref.get('order_no', 'N/A')}, Page No. {existing_ref.get('page_no', 'N/A')}")
        
        # Check if we're creating new reference
        if f"creating_new_ref_{allocation_type}" in st.session_state and st.session_state[f"creating_new_ref_{allocation_type}"]:
            return create_reference_form(allocation_type)
        
        return None
    else:
        return create_reference_form(allocation_type)

def create_reference_form(allocation_type):
    """Create a form for entering reference details"""
    st.markdown(f"### 📋 Enter Reference for {allocation_type}")
    
    order_no = st.text_input("Order No.:", key=f"order_no_{allocation_type}")
    page_no = st.text_input("Page No.:", key=f"page_no_{allocation_type}")
    remarks = st.text_area("Remarks (Optional):", key=f"remarks_{allocation_type}", height=100)
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("💾 Save Reference", key=f"save_ref_{allocation_type}"):
            if order_no and page_no:
                exam_key = st.session_state.current_exam_key
                if exam_key not in st.session_state.allocation_references:
                    st.session_state.allocation_references[exam_key] = {}
                
                st.session_state.allocation_references[exam_key][allocation_type] = {
                    'order_no': order_no,
                    'page_no': page_no,
                    'remarks': remarks,
                    'timestamp': datetime.now().isoformat(),
                    'allocation_type': allocation_type
                }
                
                save_all_data()
                st.success("✅ Reference saved successfully!")
                
                # Clear creating flag
                if f"creating_new_ref_{allocation_type}" in st.session_state:
                    st.session_state[f"creating_new_ref_{allocation_type}"] = False
                
                st.rerun()
                return st.session_state.allocation_references[exam_key][allocation_type]
            else:
                st.error("Please enter both Order No. and Page No.")
    
    with col2:
        if st.button("❌ Cancel", key=f"cancel_ref_{allocation_type}"):
            # Clear creating flag
            if f"creating_new_ref_{allocation_type}" in st.session_state:
                st.session_state[f"creating_new_ref_{allocation_type}"] = False
            st.rerun()
            return None
    
    return None

# ============================================================================
# REPORT GENERATION MODULE
# ============================================================================

def export_to_excel():
    """Generate comprehensive allocation report with all sheets"""
    if not st.session_state.allocation and not st.session_state.ey_allocation:
        st.warning("No allocation data available")
        return None
    
    try:
        # Create Excel writer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Get current exam data
            io_allocations = st.session_state.allocation
            ey_allocations = st.session_state.ey_allocation
            
            # Convert to DataFrames
            io_df = pd.DataFrame(io_allocations)
            ey_df = pd.DataFrame(ey_allocations)
            
            # ============================================
            # A. IO ALLOCATIONS SHEET
            # ============================================
            if not io_df.empty:
                # Merge with IO master data
                if not st.session_state.io_df.empty:
                    io_master = st.session_state.io_df.copy()
                    # Ensure NAME column exists in both
                    if 'NAME' in io_master.columns and 'IO Name' in io_df.columns:
                        # Standardize column names for merge
                        io_master.rename(columns={'NAME': 'IO Name'}, inplace=True)
                        io_merged = pd.merge(io_df, io_master, on='IO Name', how='left', suffixes=('', '_master'))
                    else:
                        io_merged = io_df.copy()
                else:
                    io_merged = io_df.copy()
                
                # Merge with Venue master data
                if not st.session_state.venue_df.empty:
                    venue_master = st.session_state.venue_df.copy()
                    # Standardize column names
                    venue_master.rename(columns={'VENUE': 'Venue', 'DATE': 'Date', 'SHIFT': 'Shift'}, inplace=True)
                    
                    # Create a unique key for merging
                    io_merged['merge_key'] = io_merged['Venue'] + '|' + io_merged['Date'] + '|' + io_merged['Shift']
                    venue_master['merge_key'] = venue_master['Venue'] + '|' + venue_master['Date'] + '|' + venue_master['Shift']
                    
                    io_merged = pd.merge(io_merged, venue_master, on='merge_key', how='left', suffixes=('', '_venue'))
                    io_merged.drop('merge_key', axis=1, inplace=True)
                
                # Write to Excel
                io_merged.to_excel(writer, sheet_name='IO Allocations', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['IO Allocations']
                apply_formatting(worksheet)
            
            # ============================================
            # B. EY ALLOCATIONS SHEET
            # ============================================
            if not ey_df.empty:
                # Write EY allocations
                ey_df.to_excel(writer, sheet_name='EY Allocations', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['EY Allocations']
                apply_formatting(worksheet)
            
            # ============================================
            # C. IO SUMMARY SHEET
            # ============================================
            if not io_df.empty:
                io_summary_data = []
                
                for io_name in io_df['IO Name'].unique():
                    io_data = io_df[io_df['IO Name'] == io_name]
                    first_row = io_data.iloc[0]
                    
                    # Get IO master details
                    io_master_info = {}
                    if not st.session_state.io_df.empty and 'NAME' in st.session_state.io_df.columns:
                        io_master_row = st.session_state.io_df[st.session_state.io_df['NAME'] == io_name]
                        if not io_master_row.empty:
                            io_master_info = io_master_row.iloc[0].to_dict()
                    
                    # Calculate statistics
                    total_days = io_data['Date'].nunique()
                    total_shifts = len(io_data)
                    total_venues = io_data['Venue'].nunique()
                    
                    # Get venues with dates
                    venue_details = []
                    for venue in io_data['Venue'].unique():
                        venue_dates = io_data[io_data['Venue'] == venue]['Date'].unique()
                        venue_dates_str = ", ".join(sorted(venue_dates))
                        
                        # Get venue address if available
                        venue_address = ""
                        if not st.session_state.venue_df.empty:
                            venue_info = st.session_state.venue_df[
                                st.session_state.venue_df['VENUE'] == venue
                            ]
                            if not venue_info.empty:
                                venue_address = venue_info.iloc[0].get('ADDRESS', '')
                        
                        venue_details.append(f"{venue} ({venue_dates_str}) - {venue_address}")
                    
                    # Calculate shift types
                    date_shift_counts = io_data.groupby('Date')['Shift'].nunique()
                    multiple_shift_days = (date_shift_counts > 1).sum()
                    single_shift_days = (date_shift_counts == 1).sum()
                    
                    # Get dates for each shift type
                    multiple_shift_dates = []
                    single_shift_dates = []
                    for date, shift_count in date_shift_counts.items():
                        if shift_count > 1:
                            multiple_shift_dates.append(date)
                        else:
                            single_shift_dates.append(date)
                    
                    # Prepare summary row
                    summary_row = {
                        'IO Name': io_name,
                        'Role': first_row.get('Role', ''),
                        'Total Venues Assigned': total_venues,
                        'Venues with Dates': "\n".join(venue_details),
                        'Total Days': total_days,
                        'Total Shifts': total_shifts,
                        'Multiple Shift Days': multiple_shift_days,
                        'Single Shift Days': single_shift_days,
                        'Multiple Shift Dates': ", ".join(multiple_shift_dates),
                        'Single Shift Dates': ", ".join(single_shift_dates),
                        'Area': io_master_info.get('AREA', ''),
                        'Designation': io_master_info.get('DESIGNATION', ''),
                        'Mobile': io_master_info.get('MOBILE', ''),
                        'Email': io_master_info.get('EMAIL', ''),
                        'Bank Name': io_master_info.get('BANK_NAME', ''),
                        'Account Number': io_master_info.get('ACCOUNT_NUMBER', ''),
                        'IFSC Code': io_master_info.get('IFSC_CODE', '')
                    }
                    
                    io_summary_data.append(summary_row)
                
                # Create DataFrame and write to Excel
                io_summary_df = pd.DataFrame(io_summary_data)
                io_summary_df.to_excel(writer, sheet_name='IO Summary', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['IO Summary']
                apply_formatting(worksheet)
            
            # ============================================
            # D. VENUE-IO SHIFTS SHEET
            # ============================================
            if not io_df.empty:
                venue_io_data = []
                
                # Get unique venues
                venues = io_df['Venue'].unique()
                
                for venue in venues:
                    venue_io_df = io_df[io_df['Venue'] == venue]
                    
                    # Get venue master info
                    venue_info = {}
                    if not st.session_state.venue_df.empty:
                        venue_master = st.session_state.venue_df[
                            st.session_state.venue_df['VENUE'] == venue
                        ]
                        if not venue_master.empty:
                            venue_info = venue_master.iloc[0].to_dict()
                    
                    # Group by IO
                    for io_name in venue_io_df['IO Name'].unique():
                        io_venue_data = venue_io_df[venue_io_df['IO Name'] == io_name]
                        first_row = io_venue_data.iloc[0]
                        
                        # Calculate statistics for this IO at this venue
                        total_days_at_venue = io_venue_data['Date'].nunique()
                        total_shifts_at_venue = len(io_venue_data)
                        
                        # Calculate shift types
                        date_shift_counts = io_venue_data.groupby('Date')['Shift'].nunique()
                        multiple_shifts = (date_shift_counts > 1).sum()
                        single_shifts = (date_shift_counts == 1).sum()
                        
                        # Get dates for each shift type
                        multiple_shift_dates = []
                        single_shift_dates = []
                        for date, shift_count in date_shift_counts.items():
                            if shift_count > 1:
                                multiple_shift_dates.append(date)
                            else:
                                single_shift_dates.append(date)
                        
                        # Get IO master info
                        io_master_info = {}
                        if not st.session_state.io_df.empty and 'NAME' in st.session_state.io_df.columns:
                            io_master_row = st.session_state.io_df[st.session_state.io_df['NAME'] == io_name]
                            if not io_master_row.empty:
                                io_master_info = io_master_row.iloc[0].to_dict()
                        
                        # Prepare row
                        venue_io_row = {
                            'Venue': venue,
                            'IO Name': io_name,
                            'Role': first_row.get('Role', ''),
                            'Total Days at Venue': total_days_at_venue,
                            'Multiple Shifts Count': multiple_shifts,
                            'Single Shifts Count': single_shifts,
                            'Multiple Shift Dates': ", ".join(multiple_shift_dates),
                            'Single Shift Dates': ", ".join(single_shift_dates),
                            'IO Area': io_master_info.get('AREA', ''),
                            'IO Designation': io_master_info.get('DESIGNATION', ''),
                            'IO Mobile': io_master_info.get('MOBILE', ''),
                            'Venue Address': venue_info.get('ADDRESS', ''),
                            'Centre Code': venue_info.get('CENTRE_CODE', ''),
                            'Centre Name': venue_info.get('CENTRE NAME', venue),
                            'District': venue_info.get('DISTRICT', ''),
                            'Capacity': venue_info.get('CAPACITY', '')
                        }
                        
                        venue_io_data.append(venue_io_row)
                
                # Create DataFrame and write to Excel
                if venue_io_data:
                    venue_io_df = pd.DataFrame(venue_io_data)
                    venue_io_df.to_excel(writer, sheet_name='Venue-IO Shifts', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['Venue-IO Shifts']
                    apply_formatting(worksheet)
            
            # ============================================
            # E. VENUE-ROLE SUMMARY SHEET
            # ============================================
            if not io_df.empty:
                venue_role_data = []
                
                for venue in io_df['Venue'].unique():
                    venue_data = io_df[io_df['Venue'] == venue]
                    
                    for role in venue_data['Role'].unique():
                        role_data = venue_data[venue_data['Role'] == role]
                        
                        # Count assignments
                        assignments_count = len(role_data)
                        
                        # Get unique dates
                        unique_dates = sorted(role_data['Date'].unique())
                        dates_str = ", ".join(unique_dates)
                        
                        # Get unique IOs
                        unique_ios = sorted(role_data['IO Name'].unique())
                        ios_str = ", ".join(unique_ios)
                        
                        venue_role_row = {
                            'Venue': venue,
                            'Role': role,
                            'Assignments Count': assignments_count,
                            'Dates': dates_str,
                            'IOs Assigned': ios_str
                        }
                        
                        venue_role_data.append(venue_role_row)
                
                # Create DataFrame and write to Excel
                if venue_role_data:
                    venue_role_df = pd.DataFrame(venue_role_data)
                    venue_role_df.to_excel(writer, sheet_name='Venue-Role Summary', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['Venue-Role Summary']
                    apply_formatting(worksheet)
            
            # ============================================
            # F. DATE SUMMARY SHEET
            # ============================================
            if not io_df.empty:
                date_summary_data = []
                
                for date in io_df['Date'].unique():
                    date_data = io_df[io_df['Date'] == date]
                    
                    unique_venues = date_data['Venue'].nunique()
                    unique_ios = date_data['IO Name'].nunique()
                    total_shifts = len(date_data)
                    
                    # Get venue list
                    venues_list = sorted(date_data['Venue'].unique())
                    venues_str = ", ".join(venues_list)
                    
                    # Get IO list
                    ios_list = sorted(date_data['IO Name'].unique())
                    ios_str = ", ".join(ios_list)
                    
                    date_summary_row = {
                        'Date': date,
                        'Unique Venues': unique_venues,
                        'Unique IOs': unique_ios,
                        'Total Shifts': total_shifts,
                        'Venues': venues_str,
                        'IOs': ios_str
                    }
                    
                    date_summary_data.append(date_summary_row)
                
                # Create DataFrame and write to Excel
                if date_summary_data:
                    date_summary_df = pd.DataFrame(date_summary_data)
                    date_summary_df.to_excel(writer, sheet_name='Date Summary', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['Date Summary']
                    apply_formatting(worksheet)
            
            # ============================================
            # G. EY SUMMARY SHEET
            # ============================================
            if not ey_df.empty:
                ey_summary_data = []
                
                for ey_person in ey_df['EY Personnel'].unique():
                    ey_person_data = ey_df[ey_df['EY Personnel'] == ey_person]
                    first_row = ey_person_data.iloc[0]
                    
                    # Calculate statistics
                    total_venues = ey_person_data['Venue'].nunique()
                    total_days = ey_person_data['Date'].nunique()
                    total_shifts = len(ey_person_data)
                    
                    # Get venues list
                    venues_list = sorted(ey_person_data['Venue'].unique())
                    venues_str = ", ".join(venues_list)
                    
                    # Get dates list
                    dates_list = sorted(ey_person_data['Date'].unique())
                    dates_str = ", ".join(dates_list)
                    
                    # Calculate amount
                    rate = first_row.get('Rate (₹)', st.session_state.remuneration_rates['ey_personnel'])
                    total_amount = total_days * rate
                    
                    ey_summary_row = {
                        'EY Personnel': ey_person,
                        'Venues': venues_str,
                        'Total Days': total_days,
                        'Total Shifts': total_shifts,
                        'Rate per Day (₹)': rate,
                        'Total Amount (₹)': total_amount,
                        'Dates': dates_str,
                        'Mobile': first_row.get('Mobile', ''),
                        'Email': first_row.get('Email', ''),
                        'ID Number': first_row.get('ID Number', ''),
                        'Designation': first_row.get('Designation', '')
                    }
                    
                    ey_summary_data.append(ey_summary_row)
                
                # Create DataFrame and write to Excel
                if ey_summary_data:
                    ey_summary_df = pd.DataFrame(ey_summary_data)
                    ey_summary_df.to_excel(writer, sheet_name='EY Summary', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['EY Summary']
                    apply_formatting(worksheet)
            
            # ============================================
            # H. DELETED RECORDS SHEET
            # ============================================
            if st.session_state.deleted_records:
                deleted_df = pd.DataFrame(st.session_state.deleted_records)
                deleted_df.to_excel(writer, sheet_name='Deleted Records', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['Deleted Records']
                apply_formatting(worksheet)
            
            # ============================================
            # I. RATES SHEET
            # ============================================
            rates_data = {
                'Category': ['Multiple Shifts', 'Single Shift', 'Mock Test', 'EY Personnel'],
                'Amount (₹)': [
                    st.session_state.remuneration_rates['multiple_shifts'],
                    st.session_state.remuneration_rates['single_shift'],
                    st.session_state.remuneration_rates['mock_test'],
                    st.session_state.remuneration_rates['ey_personnel']
                ],
                'Description': [
                    'For assignments with multiple shifts on same day',
                    'For assignments with single shift',
                    'For mock test assignments',
                    'Daily rate for EY personnel'
                ]
            }
            
            rates_df = pd.DataFrame(rates_data)
            rates_df.to_excel(writer, sheet_name='Rates', index=False)
            
            # Apply formatting
            worksheet = writer.sheets['Rates']
            apply_formatting(worksheet)
            
            # ============================================
            # Auto-adjust column widths
            # ============================================
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
    
    except Exception as e:
        st.error(f"Error generating report: {str(e)}")
        logging.error(f"Report generation error: {str(e)}")
        return None

def export_remuneration_report():
    """Generate comprehensive remuneration report with all sheets"""
    if not st.session_state.allocation and not st.session_state.ey_allocation:
        st.warning("No allocation data available")
        return None
    
    try:
        # Create Excel writer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Get current exam data
            io_allocations = st.session_state.allocation
            ey_allocations = st.session_state.ey_allocation
            
            # Convert to DataFrames
            io_df = pd.DataFrame(io_allocations)
            ey_df = pd.DataFrame(ey_allocations)
            
            # ============================================
            # A. IO DETAILED REPORT SHEET
            # ============================================
            if not io_df.empty:
                io_detailed_data = []
                
                for (io_name, date), group in io_df.groupby(['IO Name', 'Date']):
                    first_row = group.iloc[0]
                    
                    # Get venues for this day
                    venues = sorted(group['Venue'].unique())
                    venues_str = ", ".join(venues)
                    
                    # Calculate shift details
                    total_shifts = len(group)
                    shift_list = sorted(group['Shift'].unique())
                    shift_details = ", ".join(shift_list)
                    
                    # Determine shift type and amount
                    mock_test = any(group['Mock Test'])
                    if mock_test:
                        shift_type = "Mock Test"
                        amount = st.session_state.remuneration_rates['mock_test']
                    else:
                        if total_shifts > 1:
                            shift_type = "Multiple Shifts"
                            amount = st.session_state.remuneration_rates['multiple_shifts']
                        else:
                            shift_type = "Single Shift"
                            amount = st.session_state.remuneration_rates['single_shift']
                    
                    # Get reference info
                    order_no = first_row.get('Order No.', '')
                    page_no = first_row.get('Page No.', '')
                    
                    io_detailed_row = {
                        'IO Name': io_name,
                        'Venues': venues_str,
                        'Role': first_row.get('Role', ''),
                        'Date': date,
                        'Total Shifts': total_shifts,
                        'Shift Type': shift_type,
                        'Shift Details': shift_details,
                        'Mock Test': 'Yes' if mock_test else 'No',
                        'Amount (₹)': amount,
                        'Order No.': order_no,
                        'Page No.': page_no
                    }
                    
                    io_detailed_data.append(io_detailed_row)
                
                # Create DataFrame and write to Excel
                if io_detailed_data:
                    io_detailed_df = pd.DataFrame(io_detailed_data)
                    io_detailed_df.to_excel(writer, sheet_name='IO Detailed Report', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['IO Detailed Report']
                    apply_formatting(worksheet)
            
            # ============================================
            # B. IO SUMMARY SHEET
            # ============================================
            if not io_df.empty:
                io_summary_data = []
                
                for io_name in io_df['IO Name'].unique():
                    io_data = io_df[io_df['IO Name'] == io_name]
                    first_row = io_data.iloc[0]
                    
                    # Get IO master details
                    io_master_info = {}
                    if not st.session_state.io_df.empty and 'NAME' in st.session_state.io_df.columns:
                        io_master_row = st.session_state.io_df[st.session_state.io_df['NAME'] == io_name]
                        if not io_master_row.empty:
                            io_master_info = io_master_row.iloc[0].to_dict()
                    
                    # Calculate statistics
                    total_days = io_data['Date'].nunique()
                    total_shifts = len(io_data)
                    total_venues = io_data['Venue'].nunique()
                    
                    # Get venues list
                    venues_list = sorted(io_data['Venue'].unique())
                    venues_str = ", ".join(venues_list)
                    
                    # Calculate shift types and amounts
                    mock_days = io_data[io_data['Mock Test']]['Date'].nunique()
                    exam_days = total_days - mock_days
                    
                    date_shift_counts = io_data.groupby('Date')['Shift'].nunique()
                    multiple_shift_days = (date_shift_counts > 1).sum()
                    single_shift_days = (date_shift_counts == 1).sum()
                    
                    # Get dates for each category
                    mock_dates = sorted(io_data[io_data['Mock Test']]['Date'].unique())
                    multiple_shift_dates = []
                    single_shift_dates = []
                    
                    for date, shift_count in date_shift_counts.items():
                        if shift_count > 1 and date not in mock_dates:
                            multiple_shift_dates.append(date)
                        elif shift_count == 1 and date not in mock_dates:
                            single_shift_dates.append(date)
                    
                    # Calculate amounts
                    mock_amount = mock_days * st.session_state.remuneration_rates['mock_test']
                    multiple_amount = multiple_shift_days * st.session_state.remuneration_rates['multiple_shifts']
                    single_amount = single_shift_days * st.session_state.remuneration_rates['single_shift']
                    total_amount = mock_amount + multiple_amount + single_amount
                    
                    # Prepare summary row
                    summary_row = {
                        'IO Name': io_name,
                        'Role': first_row.get('Role', ''),
                        'Venues': venues_str,
                        'Total Amount (₹)': total_amount,
                        'Total Shifts': total_shifts,
                        'Total Days': total_days,
                        'Mock Days': mock_days,
                        'Exam Days': exam_days,
                        'Multiple Shift Days': multiple_shift_days,
                        'Single Shift Days': single_shift_days,
                        'Mock Dates': ", ".join(mock_dates),
                        'Multiple Shift Dates': ", ".join(multiple_shift_dates),
                        'Single Shift Dates': ", ".join(single_shift_dates),
                        'Mock Amount (₹)': mock_amount,
                        'Multiple Shift Amount (₹)': multiple_amount,
                        'Single Shift Amount (₹)': single_amount,
                        'Area': io_master_info.get('AREA', ''),
                        'Designation': io_master_info.get('DESIGNATION', ''),
                        'Mobile': io_master_info.get('MOBILE', ''),
                        'Email': io_master_info.get('EMAIL', ''),
                        'Bank Name': io_master_info.get('BANK_NAME', ''),
                        'Account Number': io_master_info.get('ACCOUNT_NUMBER', ''),
                        'IFSC Code': io_master_info.get('IFSC_CODE', '')
                    }
                    
                    io_summary_data.append(summary_row)
                
                # Create DataFrame and write to Excel
                if io_summary_data:
                    io_summary_df = pd.DataFrame(io_summary_data)
                    io_summary_df.to_excel(writer, sheet_name='IO Summary', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['IO Summary']
                    apply_formatting(worksheet)
            
            # ============================================
            # C. EY PERSONNEL REPORT SHEET
            # ============================================
            if not ey_df.empty:
                ey_report_data = []
                
                for (ey_person, date), group in ey_df.groupby(['EY Personnel', 'Date']):
                    first_row = group.iloc[0]
                    
                    # Get venues for this day
                    venues = sorted(group['Venue'].unique())
                    venues_str = ", ".join(venues)
                    
                    # Calculate shift details
                    total_shifts = len(group)
                    shift_list = sorted(group['Shift'].unique())
                    shift_details = ", ".join(shift_list)
                    
                    # Get rate and amount
                    rate = first_row.get('Rate (₹)', st.session_state.remuneration_rates['ey_personnel'])
                    amount = rate  # Per day rate
                    
                    # Get reference info
                    order_no = first_row.get('Order No.', '')
                    page_no = first_row.get('Page No.', '')
                    
                    ey_report_row = {
                        'EY Personnel': ey_person,
                        'Venues': venues_str,
                        'Date': date,
                        'Total Shifts': total_shifts,
                        'Shift Details': shift_details,
                        'Mock Test': 'No',
                        'Amount (₹)': amount,
                        'Rate Type': 'Per Day',
                        'Order No.': order_no,
                        'Page No.': page_no
                    }
                    
                    ey_report_data.append(ey_report_row)
                
                # Create DataFrame and write to Excel
                if ey_report_data:
                    ey_report_df = pd.DataFrame(ey_report_data)
                    ey_report_df.to_excel(writer, sheet_name='EY Personnel Report', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['EY Personnel Report']
                    apply_formatting(worksheet)
            
            # ============================================
            # D. EY SUMMARY SHEET
            # ============================================
            if not ey_df.empty:
                ey_summary_data = []
                
                for ey_person in ey_df['EY Personnel'].unique():
                    ey_person_data = ey_df[ey_df['EY Personnel'] == ey_person]
                    first_row = ey_person_data.iloc[0]
                    
                    # Calculate statistics
                    total_venues = ey_person_data['Venue'].nunique()
                    total_days = ey_person_data['Date'].nunique()
                    total_shifts = len(ey_person_data)
                    
                    # Get venues list
                    venues_list = sorted(ey_person_data['Venue'].unique())
                    venues_str = ", ".join(venues_list)
                    
                    # Get dates list
                    dates_list = sorted(ey_person_data['Date'].unique())
                    dates_str = ", ".join(dates_list)
                    
                    # Calculate amount
                    rate = first_row.get('Rate (₹)', st.session_state.remuneration_rates['ey_personnel'])
                    total_amount = total_days * rate
                    
                    # Get reference info
                    order_no = first_row.get('Order No.', '')
                    page_no = first_row.get('Page No.', '')
                    
                    ey_summary_row = {
                        'EY Personnel': ey_person,
                        'Venues': venues_str,
                        'Total Amount (₹)': total_amount,
                        'Total Days': total_days,
                        'Total Shifts': total_shifts,
                        'Order No.': order_no,
                        'Page No.': page_no,
                        'Rate per Day (₹)': rate,
                        'Dates': dates_str
                    }
                    
                    ey_summary_data.append(ey_summary_row)
                
                # Create DataFrame and write to Excel
                if ey_summary_data:
                    ey_summary_df = pd.DataFrame(ey_summary_data)
                    ey_summary_df.to_excel(writer, sheet_name='EY Summary', index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets['EY Summary']
                    apply_formatting(worksheet)
            
            # ============================================
            # E. DELETED RECORDS SHEET
            # ============================================
            if st.session_state.deleted_records:
                deleted_df = pd.DataFrame(st.session_state.deleted_records)
                deleted_df.to_excel(writer, sheet_name='Deleted Records', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['Deleted Records']
                apply_formatting(worksheet)
            
            # ============================================
            # F. RATES SHEET
            # ============================================
            rates_data = {
                'Category': ['Multiple Shifts', 'Single Shift', 'Mock Test', 'EY Personnel'],
                'Amount (₹)': [
                    st.session_state.remuneration_rates['multiple_shifts'],
                    st.session_state.remuneration_rates['single_shift'],
                    st.session_state.remuneration_rates['mock_test'],
                    st.session_state.remuneration_rates['ey_personnel']
                ],
                'Description': [
                    'For assignments with multiple shifts on same day',
                    'For assignments with single shift',
                    'For mock test assignments',
                    'Daily rate for EY personnel'
                ]
            }
            
            rates_df = pd.DataFrame(rates_data)
            rates_df.to_excel(writer, sheet_name='Rates', index=False)
            
            # Apply formatting
            worksheet = writer.sheets['Rates']
            apply_formatting(worksheet)
            
            # ============================================
            # Auto-adjust column widths
            # ============================================
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
    
    except Exception as e:
        st.error(f"Error generating remuneration report: {str(e)}")
        logging.error(f"Remuneration report generation error: {str(e)}")
        return None

def apply_formatting(worksheet):
    """Apply formatting to Excel worksheet"""
    # Define styles
    header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=11)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Apply data formatting
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Format numeric cells
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'

# ============================================================================
# ENHANCED DATE SELECTION COMPONENT
# ============================================================================

def create_date_selector(venue_data, selected_venue):
    """Create enhanced date selection interface with color coding"""
    if venue_data.empty:
        st.warning(f"No data found for venue: {selected_venue}")
        return []
    
    # Get unique dates
    unique_dates = sorted(venue_data['DATE'].dropna().unique())
    
    if not unique_dates:
        st.warning(f"No dates available for {selected_venue}")
        return []
    
    st.write(f"**Available dates for {selected_venue}:**")
    
    # Initialize date selection state if not exists
    if 'date_selection_state' not in st.session_state:
        st.session_state.date_selection_state = {}
    
    if 'expanded_dates' not in st.session_state:
        st.session_state.expanded_dates = {}
    
    # Get venue key for state management
    venue_key = f"{st.session_state.current_exam_key}_{selected_venue}"
    if venue_key not in st.session_state.date_selection_state:
        st.session_state.date_selection_state[venue_key] = {}
    
    if venue_key not in st.session_state.expanded_dates:
        st.session_state.expanded_dates[venue_key] = {}
    
    selected_date_shifts = []
    
    for date_str in unique_dates:
        # Get shifts for this date
        date_shifts_data = venue_data[venue_data['DATE'] == date_str]
        date_shifts = date_shifts_data['SHIFT'].unique()
        
        # Convert to strings and filter
        date_shifts = [str(shift) for shift in date_shifts if pd.notna(shift) and str(shift) != '']
        
        if not date_shifts:
            continue
        
        # Initialize date state if not exists
        date_key = f"{venue_key}_{date_str}"
        if date_key not in st.session_state.date_selection_state[venue_key]:
            st.session_state.date_selection_state[venue_key][date_key] = {
                'all_selected': False,
                'shifts': {shift: False for shift in date_shifts}
            }
        
        if date_str not in st.session_state.expanded_dates[venue_key]:
            st.session_state.expanded_dates[venue_key][date_str] = False
        
        # Get current state
        date_state = st.session_state.date_selection_state[venue_key][date_key]
        is_expanded = st.session_state.expanded_dates[venue_key][date_str]
        
        # Calculate selection status
        selected_shifts = [shift for shift, selected in date_state['shifts'].items() if selected]
        all_selected = len(selected_shifts) == len(date_shifts)
        partially_selected = len(selected_shifts) > 0 and not all_selected
        none_selected = len(selected_shifts) == 0
        
        # Determine color based on selection
        if all_selected:
            bg_color = "#4CAF50"  # Green
            border_color = "#388E3C"
            status_text = "✓ All Selected"
            emoji = "🟢"
        elif partially_selected:
            bg_color = "#FF9800"  # Orange
            border_color = "#F57C00"
            status_text = f"✓ {len(selected_shifts)}/{len(date_shifts)} Selected"
            emoji = "🟠"
        else:
            bg_color = "#FFEB3B"  # Yellow
            border_color = "#FBC02D"
            status_text = "Not Selected"
            emoji = "🟡"
        
        # Create date header with selection status
        col1, col2, col3 = st.columns([3, 2, 1])
        
        with col1:
            # Create a styled date header
            st.markdown(f"""
                <div style="
                    background-color: {bg_color};
                    color: #333;
                    padding: 10px 15px;
                    border-radius: 8px;
                    border: 2px solid {border_color};
                    margin: 5px 0;
                    cursor: pointer;
                    font-weight: bold;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                ">
                {emoji} <strong>{date_str}</strong> - {status_text}
                </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Single click to select all
            if st.button(f"🎯 Select All", key=f"select_all_{date_key}", 
                        use_container_width=True):
                # Toggle all shifts
                if all_selected:
                    # Deselect all
                    for shift in date_shifts:
                        date_state['shifts'][shift] = False
                else:
                    # Select all
                    for shift in date_shifts:
                        date_state['shifts'][shift] = True
                st.rerun()
        
        with col3:
            # Toggle expand/collapse
            expand_label = "📖 Show Shifts" if not is_expanded else "📕 Hide Shifts"
            if st.button(expand_label, key=f"expand_{date_key}", 
                        use_container_width=True):
                st.session_state.expanded_dates[venue_key][date_str] = not is_expanded
                st.rerun()
        
        # Show shift selection if expanded
        if is_expanded:
            st.markdown("**Select Shifts:**")
            
            # Create columns for shifts
            shift_cols = st.columns(min(4, len(date_shifts)))
            
            for idx, shift in enumerate(sorted(date_shifts)):
                col_idx = idx % len(shift_cols)
                with shift_cols[col_idx]:
                    shift_selected = st.checkbox(
                        f"⏰ {shift}",
                        value=date_state['shifts'][shift],
                        key=f"shift_{date_key}_{shift}"
                    )
                    date_state['shifts'][shift] = shift_selected
            
            st.markdown("---")
        
        # Add selected shifts to result
        for shift, selected in date_state['shifts'].items():
            if selected:
                selected_date_shifts.append({
                    'date': date_str,
                    'shift': shift,
                    'is_mock': False
                })
    
    return selected_date_shifts

def create_ey_date_selector(venue_data, venue_name):
    """Create enhanced date selection interface for EY allocations"""
    if venue_data.empty:
        return []
    
    # Get unique dates
    unique_dates = sorted(venue_data['DATE'].dropna().unique())
    
    if not unique_dates:
        return []
    
    # Initialize EY date selection state if not exists
    if 'ey_date_selection_state' not in st.session_state:
        st.session_state.ey_date_selection_state = {}
    
    if 'ey_expanded_dates' not in st.session_state:
        st.session_state.ey_expanded_dates = {}
    
    # Get venue key for state management
    venue_key = f"{st.session_state.current_exam_key}_{venue_name}"
    if venue_key not in st.session_state.ey_date_selection_state:
        st.session_state.ey_date_selection_state[venue_key] = {}
    
    if venue_key not in st.session_state.ey_expanded_dates:
        st.session_state.ey_expanded_dates[venue_key] = {}
    
    selected_date_shifts = []
    
    for date_str in unique_dates:
        # Get shifts for this date
        date_shifts_data = venue_data[venue_data['DATE'] == date_str]
        date_shifts = date_shifts_data['SHIFT'].unique()
        
        # Convert to strings and filter
        date_shifts = [str(shift) for shift in date_shifts if pd.notna(shift) and str(shift) != '']
        
        if not date_shifts:
            continue
        
        # Initialize date state if not exists
        date_key = f"{venue_key}_{date_str}"
        if date_key not in st.session_state.ey_date_selection_state[venue_key]:
            st.session_state.ey_date_selection_state[venue_key][date_key] = {
                'all_selected': False,
                'shifts': {shift: False for shift in date_shifts}
            }
        
        if date_str not in st.session_state.ey_expanded_dates[venue_key]:
            st.session_state.ey_expanded_dates[venue_key][date_str] = False
        
        # Get current state
        date_state = st.session_state.ey_date_selection_state[venue_key][date_key]
        is_expanded = st.session_state.ey_expanded_dates[venue_key][date_str]
        
        # Calculate selection status
        selected_shifts = [shift for shift, selected in date_state['shifts'].items() if selected]
        all_selected = len(selected_shifts) == len(date_shifts)
        partially_selected = len(selected_shifts) > 0 and not all_selected
        none_selected = len(selected_shifts) == 0
        
        # Determine color based on selection
        if all_selected:
            bg_color = "#4CAF50"  # Green
            border_color = "#388E3C"
            status_text = "✓ All Selected"
            emoji = "🟢"
        elif partially_selected:
            bg_color = "#FF9800"  # Orange
            border_color = "#F57C00"
            status_text = f"✓ {len(selected_shifts)}/{len(date_shifts)} Selected"
            emoji = "🟠"
        else:
            bg_color = "#FFEB3B"  # Yellow
            border_color = "#FBC02D"
            status_text = "Not Selected"
            emoji = "🟡"
        
        # Create date header with selection status
        col1, col2, col3 = st.columns([3, 2, 1])
        
        with col1:
            # Create a styled date header
            st.markdown(f"""
                <div style="
                    background-color: {bg_color};
                    color: #333;
                    padding: 10px 15px;
                    border-radius: 8px;
                    border: 2px solid {border_color};
                    margin: 5px 0;
                    cursor: pointer;
                    font-weight: bold;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                ">
                {emoji} <strong>{date_str}</strong> - {status_text}
                </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Single click to select all
            if st.button(f"🎯 Select All", key=f"ey_select_all_{date_key}", 
                        use_container_width=True):
                # Toggle all shifts
                if all_selected:
                    # Deselect all
                    for shift in date_shifts:
                        date_state['shifts'][shift] = False
                else:
                    # Select all
                    for shift in date_shifts:
                        date_state['shifts'][shift] = True
                st.rerun()
        
        with col3:
            # Toggle expand/collapse
            expand_label = "📖 Show Shifts" if not is_expanded else "📕 Hide Shifts"
            if st.button(expand_label, key=f"ey_expand_{date_key}", 
                        use_container_width=True):
                st.session_state.ey_expanded_dates[venue_key][date_str] = not is_expanded
                st.rerun()
        
        # Show shift selection if expanded
        if is_expanded:
            st.markdown("**Select Shifts:**")
            
            # Create columns for shifts
            shift_cols = st.columns(min(4, len(date_shifts)))
            
            for idx, shift in enumerate(sorted(date_shifts)):
                col_idx = idx % len(shift_cols)
                with shift_cols[col_idx]:
                    shift_selected = st.checkbox(
                        f"⏰ {shift}",
                        value=date_state['shifts'][shift],
                        key=f"ey_shift_{date_key}_{shift}"
                    )
                    date_state['shifts'][shift] = shift_selected
            
            st.markdown("---")
        
        # Add selected shifts to result
        for shift, selected in date_state['shifts'].items():
            if selected:
                selected_date_shifts.append({
                    'venue': venue_name,
                    'date': date_str,
                    'shift': shift,
                    'is_mock': False
                })
    
    return selected_date_shifts

# ============================================================================
# DASHBOARD MODULE
# ============================================================================

def show_dashboard():
    """Display main dashboard"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>📊 SYSTEM DASHBOARD</h1>
            <p>Comprehensive Overview of Allocation System</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Quick Stats Row
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
            <div style='background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); text-align: center;'>
                <h3 style='color: #4169e1;'>👨‍💼 IO Allocations</h3>
                <h1 style='color: #2c3e50;'>{len(st.session_state.allocation)}</h1>
                <p style='color: #7f8c8d;'>Active Entries</p>
            </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
            <div style='background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); text-align: center;'>
                <h3 style='color: #9370db;'>👁️ EY Allocations</h3>
                <h1 style='color: #2c3e50;'>{len(st.session_state.ey_allocation)}</h1>
                <p style='color: #7f8c8d;'>Active Entries</p>
            </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
            <div style='background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); text-align: center;'>
                <h3 style='color: #20b2aa;'>📚 Total Exams</h3>
                <h1 style='color: #2c3e50;'>{len(st.session_state.exam_data)}</h1>
                <p style='color: #7f8c8d;'>Created Exams</p>
            </div>
        """, unsafe_allow_html=True)
    
    with col4:
        current_exam = st.session_state.current_exam_key or "Not Selected"
        st.markdown(f"""
            <div style='background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); text-align: center;'>
                <h3 style='color: #ff8c00;'>🎯 Active Exam</h3>
                <h4 style='color: #2c3e50;'>{current_exam[:20]}{'...' if len(current_exam) > 20 else ''}</h4>
                <p style='color: #7f8c8d;'>Currently Selected</p>
            </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Recent Activity and Quick Actions
    col5, col6 = st.columns([2, 1])
    
    with col5:
        st.markdown("### 📈 Recent Activity")
        
        if st.session_state.allocation or st.session_state.ey_allocation:
            # Show recent IO allocations
            if st.session_state.allocation:
                recent_io = st.session_state.allocation[-5:] if len(st.session_state.allocation) >= 5 else st.session_state.allocation
                if recent_io:
                    io_df = pd.DataFrame(recent_io)[['IO Name', 'Venue', 'Date', 'Shift', 'Role']]
                    st.dataframe(io_df, use_container_width=True, hide_index=True)
            else:
                st.info("No recent IO allocations")
            
            # Show recent EY allocations
            if st.session_state.ey_allocation:
                st.markdown("#### Recent EY Allocations")
                recent_ey = st.session_state.ey_allocation[-5:] if len(st.session_state.ey_allocation) >= 5 else st.session_state.ey_allocation
                if recent_ey:
                    ey_df = pd.DataFrame(recent_ey)[['EY Personnel', 'Venue', 'Date', 'Shift']]
                    st.dataframe(ey_df, use_container_width=True, hide_index=True)
        else:
            st.info("No recent activity to display")
    
    with col6:
        st.markdown("### ⚡ Quick Actions")
        
        if st.button("📥 Load Default Data", use_container_width=True):
            load_default_master_data()
            st.success("Default data loaded!")
            st.rerun()
        
        if st.button("🔄 Refresh Data", use_container_width=True):
            load_all_data()
            st.success("Data refreshed!")
            st.rerun()
        
        if st.button("💾 Create Backup", use_container_width=True):
            backup_file = create_backup("manual_backup")
            if backup_file:
                st.success(f"Backup created: {backup_file.name}")
            else:
                st.error("Failed to create backup")
        
        if st.button("📊 View All Reports", use_container_width=True):
            st.session_state.menu = "Reports"
            st.rerun()
        
        st.markdown("---")
        st.markdown("### ℹ️ System Status")
        st.info(f"**Last Updated:** {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        
        # Check data integrity
        total_records = len(st.session_state.allocation) + len(st.session_state.ey_allocation)
        if total_records > 0:
            st.success(f"✅ System is operational with {total_records} records")
        else:
            st.warning("⚠️ System is ready but no records found")

# ============================================================================
# EXAM MANAGEMENT MODULE - FIXED
# ============================================================================

def show_exam_management():
    """Display exam management interface"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #20b2aa 0%, #3cb371 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>📝 EXAM MANAGEMENT</h1>
            <p>Create, Load, and Manage Examination Data</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Two column layout
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Create/Update Exam
        st.markdown("### 🆕 Create / Update Exam")
        
        with st.container():
            exam_name = st.text_input("Exam Name:", 
                                    value=st.session_state.exam_name,
                                    placeholder="e.g., Combined Graduate Level Examination")
            
            current_year = datetime.now().year
            year_options = [str(y) for y in range(current_year - 5, current_year + 3)]
            exam_year = st.selectbox("Exam Year:", 
                                   year_options,
                                   index=year_options.index(str(current_year)) if str(current_year) in year_options else 0)
            
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✅ Create/Update Exam", use_container_width=True):
                    if exam_name.strip():
                        exam_key = f"{exam_name.strip()} - {exam_year}"
                        
                        # Set current exam
                        st.session_state.current_exam_key = exam_key
                        st.session_state.exam_name = exam_name.strip()
                        st.session_state.exam_year = exam_year
                        
                        # Initialize if new exam
                        if exam_key not in st.session_state.exam_data:
                            st.session_state.exam_data[exam_key] = {
                                'io_allocations': [],
                                'ey_allocations': []
                            }
                            st.success(f"🎉 New exam '{exam_key}' created!")
                        else:
                            # Load existing allocations
                            exam_data = st.session_state.exam_data[exam_key]
                            if isinstance(exam_data, dict):
                                st.session_state.allocation = exam_data.get('io_allocations', [])
                                st.session_state.ey_allocation = exam_data.get('ey_allocations', [])
                            else:
                                st.session_state.allocation = exam_data
                                st.session_state.ey_allocation = []
                            
                            st.success(f"📂 Exam '{exam_key}' loaded!")
                        
                        save_all_data()
                        st.rerun()
                    else:
                        st.error("Please enter an exam name")
            
            with col_b:
                if st.button("🔄 Load Default Data", use_container_width=True):
                    load_default_master_data()
    
    with col2:
        # Select Existing Exam
        st.markdown("### 📂 Select Existing Exam")
        
        exams = sorted(st.session_state.exam_data.keys())
        if exams:
            selected_exam = st.selectbox("Choose Exam:", 
                                       exams,
                                       index=exams.index(st.session_state.current_exam_key) 
                                       if st.session_state.current_exam_key in exams else 0)
            
            if st.button("📥 Load Selected Exam", use_container_width=True):
                st.session_state.current_exam_key = selected_exam
                
                # Load exam data
                exam_data = st.session_state.exam_data[selected_exam]
                if isinstance(exam_data, dict):
                    st.session_state.allocation = exam_data.get('io_allocations', [])
                    st.session_state.ey_allocation = exam_data.get('ey_allocations', [])
                else:
                    st.session_state.allocation = exam_data
                    st.session_state.ey_allocation = []
                
                # Parse exam name and year
                if " - " in selected_exam:
                    name, year = selected_exam.split(" - ", 1)
                    st.session_state.exam_name = name
                    st.session_state.exam_year = year
                
                st.success(f"✅ Exam '{selected_exam}' loaded successfully!")
                st.rerun()
            
            # Delete exam option
            with st.expander("⚠️ Delete Exam", expanded=False):
                st.warning("This action cannot be undone!")
                if st.checkbox("I understand this will delete ALL data for this exam"):
                    if st.button("🗑️ Delete Selected Exam", type="secondary"):
                        # Create backup first
                        backup_file = create_backup(f"pre_delete_{selected_exam}")
                        
                        # Delete exam
                        del st.session_state.exam_data[selected_exam]
                        
                        # Clear if current exam
                        if st.session_state.current_exam_key == selected_exam:
                            st.session_state.allocation = []
                            st.session_state.ey_allocation = []
                            st.session_state.current_exam_key = ""
                        
                        save_all_data()
                        
                        st.success(f"✅ Exam deleted. Backup created: {backup_file.name if backup_file else 'N/A'}")
                        st.rerun()
        else:
            st.info("No exams available. Create a new exam first.")
    
    st.markdown("---")
    
    # Backup & Restore Section
    st.markdown("### 💾 Backup & Restore System")
    
    col3, col4 = st.columns(2)
    
    with col3:
        st.markdown("#### Create Backup")
        backup_desc = st.text_input("Backup Description (Optional):")
        
        if st.button("🔒 Create System Backup", use_container_width=True):
            backup_file = create_backup(backup_desc)
            if backup_file:
                st.success(f"✅ Backup created: {backup_file.name}")
            else:
                st.error("❌ Failed to create backup")
    
    with col4:
        st.markdown("#### Restore Backup")
        
        # List available backups
        backup_files = list(BACKUP_DIR.glob("*.json"))
        if backup_files:
            backup_options = [f"{f.name} ({f.stat().st_size/1024:.1f} KB)" for f in sorted(backup_files, reverse=True)]
            selected_backup = st.selectbox("Select Backup:", backup_options)
            
            if st.button("🔄 Restore from Backup", type="secondary", use_container_width=True):
                # Extract filename
                backup_filename = selected_backup.split(" (")[0]
                backup_file = BACKUP_DIR / backup_filename
                
                if st.checkbox("Confirm restore (this will overwrite current data)"):
                    if restore_from_backup(backup_file):
                        st.success("✅ Backup restored successfully!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to restore backup")
        else:
            st.info("No backup files available")
    
    # Data Management
    st.markdown("---")
    st.markdown("### 🗃️ Data Management")
    
    col5, col6, col7 = st.columns(3)
    
    with col5:
        if st.button("📤 Export All Data", use_container_width=True):
            if st.session_state.exam_data:
                # Create comprehensive export
                export_data = {
                    'exam_data': st.session_state.exam_data,
                    'allocation_references': st.session_state.allocation_references,
                    'remuneration_rates': st.session_state.remuneration_rates,
                    'deleted_records': st.session_state.deleted_records,
                    'export_timestamp': datetime.now().isoformat()
                }
                
                # Convert to JSON for download
                json_str = json.dumps(export_data, indent=4, default=str)
                st.download_button(
                    label="⬇️ Download JSON Export",
                    data=json_str,
                    file_name=f"ssc_export_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json",
                    use_container_width=True
                )
            else:
                st.warning("No data to export")
    
    with col6:
        if st.button("🧹 Clear All Data", type="secondary", use_container_width=True):
            st.warning("⚠️ This will delete ALL data including backups!")
            if st.checkbox("I confirm I want to delete ALL data"):
                # Clear all data
                st.session_state.exam_data = {}
                st.session_state.allocation = []
                st.session_state.ey_allocation = []
                st.session_state.allocation_references = {}
                st.session_state.deleted_records = []
                st.session_state.current_exam_key = ""
                
                # Delete files
                for file in [DATA_FILE, REFERENCE_FILE, DELETED_RECORDS_FILE, CONFIG_FILE]:
                    if file.exists():
                        file.unlink()
                
                # Delete backups
                for backup_file in BACKUP_DIR.glob("*.json"):
                    backup_file.unlink()
                
                st.success("✅ All data cleared successfully!")
                st.rerun()
    
    with col7:
        if st.button("📊 View References", use_container_width=True):
            show_allocation_references()

# ============================================================================
# CENTRE COORDINATOR MODULE - FIXED
# ============================================================================

def show_centre_coordinator():
    """Display Centre Coordinator allocation interface"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #4169e1 0%, #6ca0dc 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>👨‍💼 CENTRE COORDINATOR ALLOCATION</h1>
            <p>Allocate Centre Coordinators and Flying Squad Personnel</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Check if exam is selected
    if not st.session_state.current_exam_key:
        st.error("⚠️ Please select or create an exam first from Exam Management")
        return
    
    # Mode selection
    col_mode1, col_mode2 = st.columns(2)
    with col_mode1:
        st.session_state.mock_test_mode = st.checkbox("🎭 Mock Test Allocation Mode", 
                                                     value=st.session_state.mock_test_mode,
                                                     help="Enable for mock test allocations")
    
    with col_mode2:
        if st.checkbox("👁️ Switch to EY Personnel", 
                      value=st.session_state.ey_allocation_mode,
                      help="Switch to EY Personnel allocation"):
            st.session_state.ey_allocation_mode = True
            st.session_state.menu = "ey"
            st.rerun()
        else:
            st.session_state.ey_allocation_mode = False
    
    # Master Data Loading Section
    st.markdown("### 📁 Master Data Management")
    
    col_data1, col_data2, col_data3 = st.columns(3)
    
    with col_data1:
        if st.button("📤 Load IO Master", use_container_width=True):
            st.session_state.show_io_upload = True
    
    with col_data2:
        if st.button("📤 Load Venue List", use_container_width=True):
            st.session_state.show_venue_upload = True
    
    with col_data3:
        if st.button("📊 View Current Data", use_container_width=True):
            show_current_data_preview()
    
    # Show file uploaders if triggered
    if st.session_state.show_io_upload:
        uploaded_io = st.file_uploader("Upload Centre Coordinator Master (Excel)", 
                                      type=['xlsx', 'xls'],
                                      key="io_master_upload")
        if uploaded_io:
            try:
                st.session_state.io_df = pd.read_excel(uploaded_io)
                st.session_state.io_df.columns = [str(col).strip().upper() for col in st.session_state.io_df.columns]
                
                required_cols = ["NAME", "AREA", "CENTRE_CODE"]
                missing_cols = [col for col in required_cols if col not in st.session_state.io_df.columns]
                
                if missing_cols:
                    st.error(f"❌ Missing columns: {', '.join(missing_cols)}")
                else:
                    st.session_state.io_master_loaded = True
                    st.success(f"✅ Loaded {len(st.session_state.io_df)} IO records")
                    st.session_state.show_io_upload = False
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    if st.session_state.show_venue_upload:
        uploaded_venue = st.file_uploader("Upload Venue List (Excel)", 
                                         type=['xlsx', 'xls'],
                                         key="venue_upload")
        if uploaded_venue:
            try:
                st.session_state.venue_df = pd.read_excel(uploaded_venue)
                st.session_state.venue_df.columns = [str(col).strip().upper() for col in st.session_state.venue_df.columns]
                
                required_cols = ["VENUE", "DATE", "SHIFT"]
                missing_cols = [col for col in required_cols if col not in st.session_state.venue_df.columns]
                
                if missing_cols:
                    st.error(f"❌ Missing columns: {', '.join(missing_cols)}")
                else:
                    # Process dates - ensure they're in proper format
                    if 'DATE' in st.session_state.venue_df.columns:
                        st.session_state.venue_df['DATE'] = pd.to_datetime(
                            st.session_state.venue_df['DATE'], errors='coerce'
                        ).dt.strftime('%d-%m-%Y')
                    
                    # Clean SHIFT column - ensure it's string and has no NaN
                    if 'SHIFT' in st.session_state.venue_df.columns:
                        st.session_state.venue_df['SHIFT'] = st.session_state.venue_df['SHIFT'].astype(str).str.strip()
                        # Replace any 'nan' strings with empty string
                        st.session_state.venue_df['SHIFT'] = st.session_state.venue_df['SHIFT'].replace('nan', '')
                    
                    # Clean VENUE column
                    if 'VENUE' in st.session_state.venue_df.columns:
                        st.session_state.venue_df['VENUE'] = st.session_state.venue_df['VENUE'].astype(str).str.strip()
                    
                    # Remove any rows with empty VENUE or DATE
                    st.session_state.venue_df = st.session_state.venue_df[
                        (st.session_state.venue_df['VENUE'].notna()) & 
                        (st.session_state.venue_df['VENUE'] != '') &
                        (st.session_state.venue_df['DATE'].notna()) & 
                        (st.session_state.venue_df['DATE'] != '')
                    ]
                    
                    st.session_state.venue_master_loaded = True
                    st.success(f"✅ Loaded {len(st.session_state.venue_df)} venue records")
                    st.session_state.show_venue_upload = False
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.error("Please ensure your Excel file has the correct format with VENUE, DATE, and SHIFT columns.")
    
    # Check if we have required data
    if not st.session_state.venue_master_loaded:
        st.warning("⚠️ Please load venue list first")
        return
    
    if not st.session_state.io_master_loaded:
        st.warning("⚠️ Please load IO master data first")
        return
    
    # Selection Parameters
    st.markdown("### 🎯 Selection Parameters")
    
    col_param1, col_param2 = st.columns(2)
    
    with col_param1:
        # Venue selection
        venues = sorted(st.session_state.venue_df['VENUE'].dropna().unique())
        selected_venue = st.selectbox("Select Venue:", 
                                     venues,
                                     key="venue_select",
                                     index=0 if len(venues) == 0 else (
                                         venues.index(st.session_state.selected_venue) 
                                         if st.session_state.selected_venue in venues else 0
                                     ))
        
        if selected_venue != st.session_state.selected_venue:
            st.session_state.selected_venue = selected_venue
            st.rerun()
    
    with col_param2:
        # Role selection
        role = st.selectbox("Select Role:", 
                           ["Centre Coordinator", "Flying Squad"],
                           key="role_select",
                           index=0 if st.session_state.selected_role == "Centre Coordinator" else 1)
        
        if role != st.session_state.selected_role:
            st.session_state.selected_role = role
            st.rerun()
    
    # Date and Shift Selection
    st.markdown("### 📅 Date & Shift Selection")
    
    if st.session_state.mock_test_mode:
        # Mock test date entry
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            mock_date = st.date_input("Mock Test Date:", 
                                     value=datetime.now().date(),
                                     key="mock_date")
        with col_date2:
            mock_shift = st.selectbox("Shift:", 
                                     ["Morning", "Afternoon", "Evening"],
                                     key="mock_shift")
        
        if st.button("➕ Add Mock Test Date", key="add_mock_date"):
            date_info = {
                'date': mock_date.strftime("%d-%m-%Y"),
                'shift': mock_shift,
                'is_mock': True
            }
            if date_info not in st.session_state.selected_dates:
                st.session_state.selected_dates.append(date_info)
                st.success(f"Added {date_info['date']} ({mock_shift})")
    else:
        # Normal date selection from venue data
        venue_data = st.session_state.venue_df[
            st.session_state.venue_df['VENUE'] == selected_venue
        ]
        
        if not venue_data.empty:
            # Use enhanced date selector
            selected_date_shifts = create_date_selector(venue_data, selected_venue)
            
            if selected_date_shifts:
                st.session_state.selected_dates = selected_date_shifts
                st.info(f"✅ Selected {len(selected_date_shifts)} date-shift combinations")
            else:
                st.info("No dates selected. Please select at least one date-shift combination.")
        else:
            st.warning(f"No data found for venue: {selected_venue}")
    
    # IO Selection
    st.markdown("### 👥 Centre Coordinator Selection")
    
    # Filter IOs based on venue
    if selected_venue and not st.session_state.mock_test_mode:
        if not venue_data.empty:
            venue_row = venue_data.iloc[0]
            centre_code = str(venue_row['CENTRE_CODE']).zfill(4) if 'CENTRE_CODE' in venue_row else ''
            
            if centre_code and 'CENTRE_CODE' in st.session_state.io_df.columns:
                filtered_io = st.session_state.io_df[
                    st.session_state.io_df['CENTRE_CODE'].astype(str).str.zfill(4).str.startswith(centre_code[:4])
                ]
            else:
                filtered_io = st.session_state.io_df
        else:
            filtered_io = st.session_state.io_df
    else:
        filtered_io = st.session_state.io_df
    
    # Search functionality
    search_term = st.text_input("🔍 Search Centre Coordinator:", 
                               placeholder="Search by name or area...")
    
    if search_term:
        filtered_io = filtered_io[
            (filtered_io['NAME'].str.lower().str.contains(search_term.lower())) |
            (filtered_io['AREA'].str.lower().str.contains(search_term.lower())) |
            (filtered_io['DESIGNATION'].str.lower().str.contains(search_term.lower()))
        ]
    
    if not filtered_io.empty:
        st.write(f"**Available Centre Coordinators ({len(filtered_io)} found):**")
        
        # Display in a scrollable container
        io_container = st.container()
        
        with io_container:
            for idx, row in filtered_io.iterrows():
                name = row.get('NAME', 'N/A')
                area = row.get('AREA', 'N/A')
                designation = row.get('DESIGNATION', 'N/A')
                centre_code = row.get('CENTRE_CODE', 'N/A')
                
                # Check existing allocations
                existing_allocations = [
                    a for a in st.session_state.allocation 
                    if a['IO Name'] == name and a.get('Exam') == st.session_state.current_exam_key
                ]
                
                # Create a card for each IO
                with st.expander(f"👤 {name} ({area})", expanded=False):
                    col_info1, col_info2 = st.columns(2)
                    
                    with col_info1:
                        st.write(f"**Designation:** {designation}")
                        st.write(f"**Centre Code:** {centre_code}")
                    
                    with col_info2:
                        st.write(f"**Mobile:** {row.get('MOBILE', 'N/A')}")
                        st.write(f"**Email:** {row.get('EMAIL', 'N/A')}")
                    
                    # Show existing allocations
                    if existing_allocations:
                        st.warning(f"⚠️ Already allocated to {len(existing_allocations)} date(s)")
                        for alloc in existing_allocations[-3:]:  # Show last 3
                            st.write(f"- {alloc['Date']} {alloc['Shift']} at {alloc['Venue']}")
                    
                    # Allocation button
                    if st.session_state.selected_dates:
                        if st.button(f"✅ Allocate {name}", key=f"alloc_btn_{idx}"):
                            # Set allocation state
                            st.session_state.current_allocation_person = name
                            st.session_state.current_allocation_area = area
                            st.session_state.current_allocation_role = role
                            st.session_state.current_allocation_type = "IO"
                            st.rerun()
                    else:
                        st.info("Select dates above to enable allocation")
    else:
        st.warning("No Centre Coordinators found matching the criteria")
    
    # Handle allocation after reference selection
    if st.session_state.current_allocation_person and st.session_state.current_allocation_type == "IO":
        # Show reference selection
        st.markdown(f"### 📋 Reference for {st.session_state.current_allocation_person}")
        ref_data = get_or_create_reference(st.session_state.current_allocation_role)
        
        if ref_data is not None:
            # Perform allocation
            allocation_count = 0
            conflicts = []
            
            for date_info in st.session_state.selected_dates:
                # Check for conflicts
                conflict, message = check_allocation_conflict(
                    st.session_state.current_allocation_person, 
                    date_info['date'], 
                    date_info['shift'], 
                    selected_venue, 
                    st.session_state.current_allocation_role, 
                    st.session_state.current_allocation_type
                )
                
                if conflict:
                    conflicts.append(message)
                    continue
                
                # Create allocation
                allocation = {
                    'Sl. No.': len(st.session_state.allocation) + allocation_count + 1,
                    'Venue': selected_venue,
                    'Date': date_info['date'],
                    'Shift': date_info['shift'],
                    'IO Name': st.session_state.current_allocation_person,
                    'Area': st.session_state.current_allocation_area,
                    'Role': st.session_state.current_allocation_role,
                    'Mock Test': date_info['is_mock'],
                    'Exam': st.session_state.current_exam_key,
                    'Order No.': ref_data['order_no'],
                    'Page No.': ref_data['page_no'],
                    'Reference Remarks': ref_data.get('remarks', ''),
                    'Timestamp': datetime.now().isoformat()
                }
                
                st.session_state.allocation.append(allocation)
                allocation_count += 1
            
            # Update exam data
            exam_key = st.session_state.current_exam_key
            if exam_key not in st.session_state.exam_data:
                st.session_state.exam_data[exam_key] = {}
            
            st.session_state.exam_data[exam_key]['io_allocations'] = st.session_state.allocation
            
            # Save data
            save_all_data()
            
            # Clear allocation state
            st.session_state.current_allocation_person = None
            st.session_state.current_allocation_area = None
            st.session_state.current_allocation_role = None
            st.session_state.current_allocation_type = None
            
            if allocation_count > 0:
                success_msg = f"✅ Allocated {st.session_state.current_allocation_person} to {allocation_count} date-shift combination(s)"
                if conflicts:
                    success_msg += f"\n\n⚠️ {len(conflicts)} conflict(s) prevented allocation"
                st.success(success_msg)
                st.rerun()
            else:
                st.error("❌ No allocations made due to conflicts")
    
    # Current Allocations Display with Deletion Options
    if st.session_state.allocation:
        st.markdown("---")
        st.markdown("### 📋 Current Allocations")
        
        alloc_df = pd.DataFrame(st.session_state.allocation)
        
        # Filter by current exam
        current_allocations = alloc_df[alloc_df['Exam'] == st.session_state.current_exam_key]
        
        if not current_allocations.empty:
            # Add checkbox column for selection
            if 'selected_for_deletion' not in current_allocations.columns:
                current_allocations = current_allocations.copy()
                current_allocations['selected_for_deletion'] = False
            
            # Display table with checkboxes
            st.markdown("**Select allocations to delete:**")
            
            # Create a form for deletion
            with st.form(key='delete_allocation_form'):
                # Create editable dataframe
                edited_df = st.data_editor(
                    current_allocations[['selected_for_deletion', 'Sl. No.', 'IO Name', 'Venue', 'Date', 'Shift', 'Role', 'Mock Test']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "selected_for_deletion": st.column_config.CheckboxColumn(
                            "Select",
                            help="Select for deletion",
                            default=False,
                        ),
                        "Sl. No.": st.column_config.NumberColumn(
                            "S.No.",
                            help="Serial number",
                            disabled=True
                        ),
                        "IO Name": st.column_config.TextColumn(
                            "IO Name",
                            help="Name of IO",
                            disabled=True
                        ),
                        "Venue": st.column_config.TextColumn(
                            "Venue",
                            help="Venue name",
                            disabled=True
                        ),
                        "Date": st.column_config.TextColumn(
                            "Date",
                            help="Allocation date",
                            disabled=True
                        ),
                        "Shift": st.column_config.TextColumn(
                            "Shift",
                            help="Shift timing",
                            disabled=True
                        ),
                        "Role": st.column_config.TextColumn(
                            "Role",
                            help="Role assigned",
                            disabled=True
                        ),
                        "Mock Test": st.column_config.CheckboxColumn(
                            "Mock Test",
                            help="Mock test allocation",
                            disabled=True
                        )
                    }
                )
                
                # Get selected indices
                selected_indices = edited_df[edited_df['selected_for_deletion']].index.tolist()
                
                col_del1, col_del2, col_del3 = st.columns(3)
                
                with col_del1:
                    delete_submitted = st.form_submit_button("🗑️ Delete Selected")
                
                with col_del2:
                    select_all = st.form_submit_button("📋 Select All")
                
                with col_del3:
                    clear_all = st.form_submit_button("🧹 Clear Selection")
                
                if select_all:
                    # Select all rows
                    for idx in range(len(current_allocations)):
                        st.session_state.selected_allocation_for_deletion.append(idx)
                    st.rerun()
                
                if clear_all:
                    # Clear all selections
                    st.session_state.selected_allocation_for_deletion = []
                    st.rerun()
                
                if delete_submitted and selected_indices:
                    # Show confirmation
                    st.warning(f"⚠️ You are about to delete {len(selected_indices)} allocation(s). This action cannot be undone!")
                    
                    if st.button(f"🔥 CONFIRM DELETE {len(selected_indices)} ALLOCATION(S)"):
                        # Convert DataFrame indices to allocation indices
                        allocation_indices = []
                        for df_idx in selected_indices:
                            if df_idx < len(current_allocations):
                                sl_no = current_allocations.iloc[df_idx]['Sl. No.']
                                # Find the allocation with this serial number
                                for idx, alloc in enumerate(st.session_state.allocation):
                                    if alloc.get('Sl. No.') == sl_no:
                                        allocation_indices.append(idx)
                                        break
                        
                        # Delete in bulk
                        deleted_count = bulk_delete_allocations(allocation_indices, "IO")
                        
                        if deleted_count > 0:
                            st.success(f"✅ Successfully deleted {deleted_count} allocation(s)")
                            st.session_state.selected_allocation_for_deletion = []
                            save_all_data()
                            st.rerun()
                        else:
                            st.error("❌ Failed to delete allocations")
            
            # Individual deletion options
            st.markdown("#### Quick Individual Deletion")
            
            for idx, alloc in enumerate(st.session_state.allocation):
                if alloc.get('Exam') == st.session_state.current_exam_key:
                    col_del_quick1, col_del_quick2, col_del_quick3, col_del_quick4 = st.columns([3, 2, 2, 1])
                    
                    with col_del_quick1:
                        st.write(f"{alloc.get('IO Name')} - {alloc.get('Venue')}")
                    
                    with col_del_quick2:
                        st.write(f"{alloc.get('Date')} {alloc.get('Shift')}")
                    
                    with col_del_quick3:
                        st.write(f"{alloc.get('Role')}")
                    
                    with col_del_quick4:
                        if st.button("🗑️", key=f"delete_single_{idx}"):
                            if delete_allocation(idx, "IO"):
                                st.success(f"Deleted allocation for {alloc.get('IO Name')}")
                                st.rerun()
                            else:
                                st.error("Failed to delete allocation")
            
            # Export options
            st.markdown("---")
            col_exp1, col_exp2, col_exp3 = st.columns(3)
            
            with col_exp1:
                if st.button("📤 Export Allocations", use_container_width=True):
                    csv = current_allocations.to_csv(index=False)
                    st.download_button(
                        label="⬇️ Download CSV",
                        data=csv,
                        file_name=f"io_allocations_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
            
            with col_exp2:
                if st.button("🗑️ Clear All Allocations", type="secondary", use_container_width=True):
                    st.warning("This will delete ALL allocations for this exam!")
                    if st.checkbox("Confirm clear all allocations for this exam"):
                        # Move to deleted records
                        for alloc in st.session_state.allocation:
                            if alloc.get('Exam') == st.session_state.current_exam_key:
                                deleted_record = {
                                    **alloc,
                                    'Deletion Timestamp': datetime.now().isoformat(),
                                    'Deletion Reason': 'Bulk clear all',
                                    'Type': 'IO'
                                }
                                st.session_state.deleted_records.append(deleted_record)
                        
                        # Clear allocations
                        st.session_state.allocation = []
                        exam_key = st.session_state.current_exam_key
                        if exam_key in st.session_state.exam_data:
                            st.session_state.exam_data[exam_key]['io_allocations'] = []
                        save_all_data()
                        st.success("All allocations cleared!")
                        st.rerun()
            
            with col_exp3:
                if st.button("📊 Generate Report", use_container_width=True):
                    generate_io_report()
        else:
            st.info("No allocations for current exam")

# ============================================================================
# EY PERSONNEL MODULE - FIXED WITH PARTIAL SHIFT SELECTION
# ============================================================================

def show_ey_personnel():
    """Display EY Personnel allocation interface"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #9370db 0%, #8a2be2 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>👁️ EY PERSONNEL ALLOCATION</h1>
            <p>Allocate External Yard/Examination Personnel</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Check if exam is selected
    if not st.session_state.current_exam_key:
        st.error("⚠️ Please select or create an exam first from Exam Management")
        return
    
    # Mode switch
    col_mode1, col_mode2 = st.columns(2)
    with col_mode1:
        if st.checkbox("👨‍💼 Switch to Centre Coordinator", 
                      help="Switch to Centre Coordinator allocation"):
            st.session_state.menu = "io"
            st.rerun()
    
    # Master Data Loading
    st.markdown("### 📁 EY Master Data")
    
    col_ey1, col_ey2 = st.columns(2)
    
    with col_ey1:
        if st.button("📤 Load EY Master", use_container_width=True):
            st.session_state.show_ey_upload = True
    
    with col_ey2:
        # EY Rate Setting
        ey_rate = st.number_input("💰 EY Rate per Day (₹):", 
                                 value=st.session_state.remuneration_rates['ey_personnel'],
                                 min_value=0, step=100)
        
        if ey_rate != st.session_state.remuneration_rates['ey_personnel']:
            st.session_state.remuneration_rates['ey_personnel'] = ey_rate
            save_all_data()
            st.success("Rate updated!")
    
    # Show EY uploader if triggered
    if st.session_state.show_ey_upload:
        uploaded_ey = st.file_uploader("Upload EY Personnel Master (Excel)", 
                                      type=['xlsx', 'xls'],
                                      key="ey_master_upload")
        if uploaded_ey:
            try:
                st.session_state.ey_df = pd.read_excel(uploaded_ey)
                st.session_state.ey_df.columns = [str(col).strip().upper() for col in st.session_state.ey_df.columns]
                
                if 'NAME' not in st.session_state.ey_df.columns:
                    st.error("❌ Missing required column: NAME")
                else:
                    # Ensure optional columns exist
                    optional_cols = ["MOBILE", "EMAIL", "ID_NUMBER", "DESIGNATION", "DEPARTMENT"]
                    for col in optional_cols:
                        if col not in st.session_state.ey_df.columns:
                            st.session_state.ey_df[col] = ""
                    
                    st.session_state.ey_master_loaded = True
                    st.success(f"✅ Loaded {len(st.session_state.ey_df)} EY personnel records")
                    st.session_state.show_ey_upload = False
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    # Check for required data
    if not st.session_state.ey_master_loaded:
        st.warning("⚠️ Please load EY master data first")
        return
    
    if not st.session_state.venue_master_loaded:
        st.warning("⚠️ Please load venue list from Centre Coordinator section")
        return
    
    # Venue Selection
    st.markdown("### 🎯 Venue Selection")
    
    venues = sorted(st.session_state.venue_df['VENUE'].dropna().unique())
    selected_venues = st.multiselect("Select Venues for EY Allocation:", 
                                    venues,
                                    default=[st.session_state.selected_venue] if st.session_state.selected_venue in venues else None)
    
    if not selected_venues:
        st.info("Select at least one venue to continue")
        return
    
    # Date Selection with Partial Shifts
    st.markdown("### 📅 Date & Shift Selection")
    
    selected_date_info = []
    
    for venue in selected_venues:
        venue_data = st.session_state.venue_df[st.session_state.venue_df['VENUE'] == venue]
        
        if not venue_data.empty:
            with st.expander(f"📅 {venue}", expanded=False):
                # Use enhanced date selector for EY
                venue_date_shifts = create_ey_date_selector(venue_data, venue)
                selected_date_info.extend(venue_date_shifts)
    
    if not selected_date_info:
        st.info("Select dates and shifts to allocate EY personnel")
        return
    
    st.info(f"✅ Selected {len(selected_date_info)} date-shift combination(s)")
    
    # EY Personnel Selection
    st.markdown("### 👥 EY Personnel Selection")
    
    # Search functionality
    search_term = st.text_input("🔍 Search EY Personnel:", 
                               placeholder="Search by name, department, or ID...")
    
    if search_term:
        filtered_ey = st.session_state.ey_df[
            (st.session_state.ey_df['NAME'].str.lower().str.contains(search_term.lower())) |
            (st.session_state.ey_df['DEPARTMENT'].str.lower().str.contains(search_term.lower())) |
            (st.session_state.ey_df['ID_NUMBER'].str.lower().str.contains(search_term.lower()))
        ]
    else:
        filtered_ey = st.session_state.ey_df
    
    if not filtered_ey.empty:
        st.write(f"**Available EY Personnel ({len(filtered_ey)} found):**")
        
        # Display EY personnel
        selected_ey = st.selectbox("Select EY Personnel:", 
                                  filtered_ey['NAME'].tolist(),
                                  key="ey_person_select")
        
        if selected_ey:
            # Show details of selected EY personnel
            ey_row = filtered_ey[filtered_ey['NAME'] == selected_ey].iloc[0]
            
            col_details1, col_details2 = st.columns(2)
            with col_details1:
                st.write(f"**ID:** {ey_row.get('ID_NUMBER', 'N/A')}")
                st.write(f"**Designation:** {ey_row.get('DESIGNATION', 'N/A')}")
            with col_details2:
                st.write(f"**Department:** {ey_row.get('DEPARTMENT', 'N/A')}")
                st.write(f"**Mobile:** {ey_row.get('MOBILE', 'N/A')}")
            
            # Allocation button
            if st.button(f"✅ Allocate {selected_ey} to Selected Dates", 
                        use_container_width=True):
                # Set allocation state
                st.session_state.current_allocation_person = selected_ey
                st.session_state.current_allocation_ey_row = ey_row.to_dict()
                st.session_state.current_allocation_type = "EY"
                st.rerun()
    else:
        st.warning("No EY personnel found matching search criteria")
    
    # Handle EY allocation after reference selection
    if st.session_state.current_allocation_person and st.session_state.current_allocation_type == "EY":
        # Show reference selection
        st.markdown(f"### 📋 Reference for {st.session_state.current_allocation_person}")
        ref_data = get_or_create_reference("EY Personnel")
        
        if ref_data is not None:
            # Perform allocation
            allocation_count = 0
            conflicts = []
            ey_row = st.session_state.current_allocation_ey_row
            
            for date_info in selected_date_info:
                # Check for conflicts
                conflict, message = check_allocation_conflict(
                    st.session_state.current_allocation_person, 
                    date_info['date'], 
                    date_info['shift'],
                    date_info['venue'], 
                    "", 
                    "EY"
                )
                
                if conflict:
                    conflicts.append(message)
                    continue
                
                # Create allocation
                allocation = {
                    'Sl. No.': len(st.session_state.ey_allocation) + allocation_count + 1,
                    'Venue': date_info['venue'],
                    'Date': date_info['date'],
                    'Shift': date_info['shift'],
                    'EY Personnel': st.session_state.current_allocation_person,
                    'Mobile': ey_row.get('MOBILE', ''),
                    'Email': ey_row.get('EMAIL', ''),
                    'ID Number': ey_row.get('ID_NUMBER', ''),
                    'Designation': ey_row.get('DESIGNATION', ''),
                    'Department': ey_row.get('DEPARTMENT', ''),
                    'Mock Test': False,
                    'Exam': st.session_state.current_exam_key,
                    'Rate (₹)': st.session_state.remuneration_rates['ey_personnel'],
                    'Order No.': ref_data['order_no'],
                    'Page No.': ref_data['page_no'],
                    'Reference Remarks': ref_data.get('remarks', ''),
                    'Timestamp': datetime.now().isoformat()
                }
                
                st.session_state.ey_allocation.append(allocation)
                allocation_count += 1
            
            # Update exam data
            exam_key = st.session_state.current_exam_key
            if exam_key not in st.session_state.exam_data:
                st.session_state.exam_data[exam_key] = {}
            
            st.session_state.exam_data[exam_key]['ey_allocations'] = st.session_state.ey_allocation
            
            # Save data
            save_all_data()
            
            # Clear allocation state
            st.session_state.current_allocation_person = None
            st.session_state.current_allocation_ey_row = None
            st.session_state.current_allocation_type = None
            
            if allocation_count > 0:
                success_msg = f"✅ Allocated {st.session_state.current_allocation_person} to {allocation_count} date-shift combinations"
                if conflicts:
                    success_msg += f"\n\n⚠️ {len(conflicts)} conflict(s) prevented allocation"
                st.success(success_msg)
                st.rerun()
            else:
                st.error("❌ No allocations made due to conflicts")
    
    # Current EY Allocations with Deletion Options
    if st.session_state.ey_allocation:
        st.markdown("---")
        st.markdown("### 📋 Current EY Allocations")
        
        ey_df = pd.DataFrame(st.session_state.ey_allocation)
        
        # Filter by current exam
        current_ey = ey_df[ey_df['Exam'] == st.session_state.current_exam_key]
        
        if not current_ey.empty:
            # Add checkbox column for selection
            if 'selected_for_deletion' not in current_ey.columns:
                current_ey = current_ey.copy()
                current_ey['selected_for_deletion'] = False
            
            # Display table with checkboxes
            st.markdown("**Select allocations to delete:**")
            
            # Create a form for deletion
            with st.form(key='delete_ey_allocation_form'):
                # Create editable dataframe
                edited_df = st.data_editor(
                    current_ey[['selected_for_deletion', 'Sl. No.', 'EY Personnel', 'Venue', 'Date', 'Shift', 'Rate (₹)']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "selected_for_deletion": st.column_config.CheckboxColumn(
                            "Select",
                            help="Select for deletion",
                            default=False,
                        ),
                        "Sl. No.": st.column_config.NumberColumn(
                            "S.No.",
                            help="Serial number",
                            disabled=True
                        ),
                        "EY Personnel": st.column_config.TextColumn(
                            "EY Personnel",
                            help="Name of EY personnel",
                            disabled=True
                        ),
                        "Venue": st.column_config.TextColumn(
                            "Venue",
                            help="Venue name",
                            disabled=True
                        ),
                        "Date": st.column_config.TextColumn(
                            "Date",
                            help="Allocation date",
                            disabled=True
                        ),
                        "Shift": st.column_config.TextColumn(
                            "Shift",
                            help="Shift timing",
                            disabled=True
                        ),
                        "Rate (₹)": st.column_config.NumberColumn(
                            "Rate",
                            help="Daily rate",
                            disabled=True,
                            format="₹%d"
                        )
                    }
                )
                
                # Get selected indices
                selected_indices = edited_df[edited_df['selected_for_deletion']].index.tolist()
                
                col_del1, col_del2, col_del3 = st.columns(3)
                
                with col_del1:
                    delete_submitted = st.form_submit_button("🗑️ Delete Selected")
                
                with col_del2:
                    select_all = st.form_submit_button("📋 Select All")
                
                with col_del3:
                    clear_all = st.form_submit_button("🧹 Clear Selection")
                
                if select_all:
                    # Select all rows
                    for idx in range(len(current_ey)):
                        st.session_state.selected_ey_allocation_for_deletion.append(idx)
                    st.rerun()
                
                if clear_all:
                    # Clear all selections
                    st.session_state.selected_ey_allocation_for_deletion = []
                    st.rerun()
                
                if delete_submitted and selected_indices:
                    # Show confirmation
                    st.warning(f"⚠️ You are about to delete {len(selected_indices)} EY allocation(s). This action cannot be undone!")
                    
                    if st.button(f"🔥 CONFIRM DELETE {len(selected_indices)} EY ALLOCATION(S)"):
                        # Convert DataFrame indices to allocation indices
                        allocation_indices = []
                        for df_idx in selected_indices:
                            if df_idx < len(current_ey):
                                sl_no = current_ey.iloc[df_idx]['Sl. No.']
                                # Find the allocation with this serial number
                                for idx, alloc in enumerate(st.session_state.ey_allocation):
                                    if alloc.get('Sl. No.') == sl_no:
                                        allocation_indices.append(idx)
                                        break
                        
                        # Delete in bulk
                        deleted_count = bulk_delete_allocations(allocation_indices, "EY")
                        
                        if deleted_count > 0:
                            st.success(f"✅ Successfully deleted {deleted_count} EY allocation(s)")
                            st.session_state.selected_ey_allocation_for_deletion = []
                            save_all_data()
                            st.rerun()
                        else:
                            st.error("❌ Failed to delete EY allocations")
            
            # Individual deletion options
            st.markdown("#### Quick Individual Deletion")
            
            for idx, alloc in enumerate(st.session_state.ey_allocation):
                if alloc.get('Exam') == st.session_state.current_exam_key:
                    col_del_quick1, col_del_quick2, col_del_quick3, col_del_quick4 = st.columns([3, 2, 2, 1])
                    
                    with col_del_quick1:
                        st.write(f"{alloc.get('EY Personnel')} - {alloc.get('Venue')}")
                    
                    with col_del_quick2:
                        st.write(f"{alloc.get('Date')} {alloc.get('Shift')}")
                    
                    with col_del_quick3:
                        st.write(f"₹{alloc.get('Rate (₹)', 0)}")
                    
                    with col_del_quick4:
                        if st.button("🗑️", key=f"delete_ey_single_{idx}"):
                            if delete_allocation(idx, "EY"):
                                st.success(f"Deleted allocation for {alloc.get('EY Personnel')}")
                                st.rerun()
                            else:
                                st.error("Failed to delete allocation")
            
            # Summary statistics
            st.markdown("---")
            unique_dates = current_ey['Date'].nunique()
            total_cost = unique_dates * st.session_state.remuneration_rates['ey_personnel']
            unique_personnel = current_ey['EY Personnel'].nunique()
            
            col_stat1, col_stat2 = st.columns(2)
            with col_stat1:
                st.metric("Total EY Personnel", unique_personnel)
            with col_stat2:
                st.metric("Estimated Cost", f"₹{total_cost:,}")
            
            # Export button
            if st.button("📤 Export EY Allocations", use_container_width=True):
                csv = current_ey.to_csv(index=False)
                st.download_button(
                    label="⬇️ Download CSV",
                    data=csv,
                    file_name=f"ey_allocations_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
        else:
            st.info("No EY allocations for current exam")

# ============================================================================
# REPORTS MODULE - ENHANCED
# ============================================================================

def show_reports():
    """Display reports and export interface"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #ff8c00 0%, #ffa500 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>📊 REPORTS & EXPORTS</h1>
            <p>Generate Comprehensive Reports and Export Data</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Tab layout for different reports
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Allocation Reports", 
        "💰 Remuneration Reports", 
        "📚 Reference Reports",
        "🗑️ Deleted Records"
    ])
    
    with tab1:
        show_allocation_reports()
    
    with tab2:
        show_remuneration_reports()
    
    with tab3:
        show_reference_reports()
    
    with tab4:
        show_deleted_records()

def show_allocation_reports():
    """Display allocation reports"""
    st.markdown("### 📋 Allocation Reports")
    
    if not st.session_state.allocation and not st.session_state.ey_allocation:
        st.info("No allocation data available")
        return
    
    col_report1, col_report2 = st.columns(2)
    
    with col_report1:
        st.markdown("#### 📊 Generate Allocation Report")
        st.write("This report includes all allocation data with multiple sheets:")
        st.markdown("""
        - **IO Allocations:** Raw allocation data
        - **EY Allocations:** EY personnel data
        - **IO Summary:** Per-IO statistics
        - **Venue-IO Shifts:** Venue-centric view
        - **Venue-Role Summary:** Counts per venue
        - **Date Summary:** Daily statistics
        - **EY Summary:** EY personnel statistics
        - **Deleted Records:** Audit trail
        - **Rates:** Current remuneration rates
        """)
        
        if st.button("📤 Generate Allocation Report", use_container_width=True):
            excel_file = export_to_excel()
            if excel_file:
                st.download_button(
                    label="⬇️ Download Allocation Report (Excel)",
                    data=excel_file,
                    file_name=f"allocation_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.error("Failed to generate report")
    
    with col_report2:
        st.markdown("#### 📈 View Data Preview")
        
        if st.session_state.allocation:
            alloc_df = pd.DataFrame(st.session_state.allocation)
            
            # Show summary statistics
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                st.metric("Total IO Allocations", len(alloc_df))
            with col_stat2:
                st.metric("Unique IOs", alloc_df['IO Name'].nunique())
            with col_stat3:
                st.metric("Total Days", alloc_df['Date'].nunique())
            
            # Show preview
            with st.expander("Preview IO Allocations"):
                st.dataframe(
                    alloc_df[['Sl. No.', 'IO Name', 'Venue', 'Date', 'Shift', 'Role', 'Mock Test']].head(10),
                    use_container_width=True,
                    hide_index=True
                )
        
        if st.session_state.ey_allocation:
            st.markdown("##### EY Allocations")
            ey_df = pd.DataFrame(st.session_state.ey_allocation)
            
            col_ey1, col_ey2 = st.columns(2)
            with col_ey1:
                st.metric("Total EY Allocations", len(ey_df))
            with col_ey2:
                st.metric("Unique EY Personnel", ey_df['EY Personnel'].nunique())
            
            with st.expander("Preview EY Allocations"):
                st.dataframe(
                    ey_df[['Sl. No.', 'EY Personnel', 'Venue', 'Date', 'Shift', 'Rate (₹)']].head(10),
                    use_container_width=True,
                    hide_index=True
                )

def show_remuneration_reports():
    """Display remuneration reports"""
    st.markdown("### 💰 Remuneration Reports")
    
    if not st.session_state.allocation and not st.session_state.ey_allocation:
        st.info("No allocation data available for remuneration calculation")
        return
    
    col_report1, col_report2 = st.columns(2)
    
    with col_report1:
        st.markdown("#### 💸 Generate Remuneration Report")
        st.write("This report includes detailed remuneration calculations:")
        st.markdown("""
        - **IO Detailed Report:** Per-day calculations
        - **IO Summary:** Aggregated IO remuneration
        - **EY Personnel Report:** EY per-day calculations
        - **EY Summary:** Aggregated EY remuneration
        - **Deleted Records:** Audit trail
        - **Rates:** Current remuneration rates
        """)
        
        if st.button("💰 Generate Remuneration Report", use_container_width=True):
            excel_file = export_remuneration_report()
            if excel_file:
                st.download_button(
                    label="⬇️ Download Remuneration Report (Excel)",
                    data=excel_file,
                    file_name=f"remuneration_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.error("Failed to generate report")
    
    with col_report2:
        st.markdown("#### 📊 Remuneration Preview")
        
        if st.session_state.allocation:
            # Calculate IO remuneration
            alloc_df = pd.DataFrame(st.session_state.allocation)
            
            # Calculate total IO remuneration
            total_io_amount = 0
            for (io_name, date), group in alloc_df.groupby(['IO Name', 'Date']):
                shifts = len(group)
                is_mock = any(group['Mock Test'])
                
                if is_mock:
                    amount = st.session_state.remuneration_rates['mock_test']
                else:
                    if shifts > 1:
                        amount = st.session_state.remuneration_rates['multiple_shifts']
                    else:
                        amount = st.session_state.remuneration_rates['single_shift']
                
                total_io_amount += amount
            
            col_rem1, col_rem2 = st.columns(2)
            with col_rem1:
                st.metric("Total IO Remuneration", f"₹{total_io_amount:,}")
            
            # Calculate EY remuneration
            if st.session_state.ey_allocation:
                ey_df = pd.DataFrame(st.session_state.ey_allocation)
                total_ey_days = ey_df['Date'].nunique()
                total_ey_amount = total_ey_days * st.session_state.remuneration_rates['ey_personnel']
                
                with col_rem2:
                    st.metric("Total EY Remuneration", f"₹{total_ey_amount:,}")
            
            # Grand total
            grand_total = total_io_amount + (total_ey_amount if 'total_ey_amount' in locals() else 0)
            st.metric("Grand Total", f"₹{grand_total:,}", delta=f"₹{grand_total:,}")

def show_reference_reports():
    """Display allocation reference reports"""
    st.markdown("### 📚 Allocation References")
    
    if not st.session_state.allocation_references:
        st.info("No allocation references available")
        return
    
    # Display references
    ref_data = []
    for exam_key, roles in st.session_state.allocation_references.items():
        for role, ref in roles.items():
            timestamp = ref.get('timestamp', '')
            if timestamp:
                try:
                    timestamp = datetime.fromisoformat(timestamp).strftime("%d-%m-%Y %H:%M")
                except:
                    pass
            
            ref_data.append({
                'Exam': exam_key,
                'Role': role,
                'Order No.': ref.get('order_no', ''),
                'Page No.': ref.get('page_no', ''),
                'Timestamp': timestamp,
                'Remarks': ref.get('remarks', '')[:50] + '...' if len(ref.get('remarks', '')) > 50 else ref.get('remarks', '')
            })
    
    if ref_data:
        ref_df = pd.DataFrame(ref_data)
        st.dataframe(ref_df, use_container_width=True)
        
        # Export
        if st.button("📤 Export References", use_container_width=True):
            csv = ref_df.to_csv(index=False)
            st.download_button(
                label="⬇️ Download CSV",
                data=csv,
                file_name=f"allocation_references_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        # Delete options
        st.markdown("---")
        st.markdown("#### 🗑️ Manage References")
        
        exams = list(st.session_state.allocation_references.keys())
        if exams:
            selected_exam = st.selectbox("Select Exam:", exams)
            
            if selected_exam and selected_exam in st.session_state.allocation_references:
                roles = list(st.session_state.allocation_references[selected_exam].keys())
                if roles:
                    selected_role = st.selectbox("Select Role:", roles)
                    
                    if st.button("🗑️ Delete Selected Reference", type="secondary"):
                        if st.checkbox(f"Confirm delete reference for {selected_role} in {selected_exam}"):
                            del st.session_state.allocation_references[selected_exam][selected_role]
                            
                            # Remove exam if no references left
                            if not st.session_state.allocation_references[selected_exam]:
                                del st.session_state.allocation_references[selected_exam]
                            
                            save_all_data()
                            st.success("Reference deleted!")
                            st.rerun()

def show_deleted_records():
    """Display deleted records"""
    st.markdown("### 🗑️ Deleted Records")
    
    if not st.session_state.deleted_records:
        st.info("No deleted records available")
        return
    
    deleted_df = pd.DataFrame(st.session_state.deleted_records)
    
    # Format timestamp
    if 'Deletion Timestamp' in deleted_df.columns:
        deleted_df['Deletion Timestamp'] = pd.to_datetime(
            deleted_df['Deletion Timestamp'], errors='coerce'
        ).dt.strftime('%d-%m-%Y %H:%M')
    
    st.dataframe(deleted_df, use_container_width=True)
    
    # Statistics
    total_deleted = len(deleted_df)
    io_deleted = len([r for r in st.session_state.deleted_records if r.get('Type') == 'IO'])
    ey_deleted = len([r for r in st.session_state.deleted_records if r.get('Type') == 'EY Personnel'])
    
    col_del1, col_del2, col_del3 = st.columns(3)
    with col_del1:
        st.metric("Total Deleted", total_deleted)
    with col_del2:
        st.metric("IO Deleted", io_deleted)
    with col_del3:
        st.metric("EY Deleted", ey_deleted)
    
    # Export
    if st.button("📤 Export Deleted Records", use_container_width=True):
        csv = deleted_df.to_csv(index=False)
        st.download_button(
            label="⬇️ Download CSV",
            data=csv,
            file_name=f"deleted_records_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    # Clear deleted records
    st.markdown("---")
    if st.button("🧹 Clear All Deleted Records", type="secondary"):
        if st.checkbox("Confirm clear all deleted records"):
            st.session_state.deleted_records = []
            save_all_data()
            st.success("Deleted records cleared!")
            st.rerun()

# ============================================================================
# SETTINGS MODULE
# ============================================================================

def show_settings():
    """Display system settings"""
    st.markdown("""
        <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%); 
                 color: white; border-radius: 10px; margin-bottom: 30px;'>
            <h1>⚙️ SYSTEM SETTINGS</h1>
            <p>Configure Application Settings and Preferences</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Tab layout
    tab1, tab2, tab3, tab4 = st.tabs([
        "💰 Remuneration Rates", 
        "🛠️ Data Management", 
        "ℹ️ System Info",
        "🆘 Help & Support"
    ])
    
    with tab1:
        show_remuneration_settings()
    
    with tab2:
        show_data_management()
    
    with tab3:
        show_system_info()
    
    with tab4:
        show_help_support()

def show_remuneration_settings():
    """Display remuneration rate settings"""
    st.markdown("### 💰 Remuneration Rates Configuration")
    
    col_rate1, col_rate2 = st.columns(2)
    
    with col_rate1:
        multiple_shifts = st.number_input(
            "Multiple Shifts (₹):",
            min_value=0,
            value=st.session_state.remuneration_rates['multiple_shifts'],
            step=50,
            help="Amount for assignments with multiple shifts on same day"
        )
        
        single_shift = st.number_input(
            "Single Shift (₹):",
            min_value=0,
            value=st.session_state.remuneration_rates['single_shift'],
            step=50,
            help="Amount for assignments with single shift"
        )
    
    with col_rate2:
        mock_test = st.number_input(
            "Mock Test (₹):",
            min_value=0,
            value=st.session_state.remuneration_rates['mock_test'],
            step=50,
            help="Amount for mock test assignments"
        )
        
        ey_personnel = st.number_input(
            "EY Personnel (₹ per day):",
            min_value=0,
            value=st.session_state.remuneration_rates['ey_personnel'],
            step=100,
            help="Daily rate for EY personnel"
        )
    
    # Save button
    if st.button("💾 Save Rates", use_container_width=True):
        st.session_state.remuneration_rates = {
            'multiple_shifts': multiple_shifts,
            'single_shift': single_shift,
            'mock_test': mock_test,
            'ey_personnel': ey_personnel
        }
        
        save_all_data()
        st.success("✅ Remuneration rates saved successfully!")
    
    # Reset to defaults
    if st.button("🔄 Reset to Defaults", type="secondary"):
        st.session_state.remuneration_rates = DEFAULT_RATES.copy()
        save_all_data()
        st.success("✅ Rates reset to defaults!")
        st.rerun()
    
    # Current rates display
    st.markdown("---")
    st.markdown("### 📊 Current Rate Summary")
    
    rates_df = pd.DataFrame([
        {'Category': 'Multiple Shifts', 'Amount (₹)': st.session_state.remuneration_rates['multiple_shifts']},
        {'Category': 'Single Shift', 'Amount (₹)': st.session_state.remuneration_rates['single_shift']},
        {'Category': 'Mock Test', 'Amount (₹)': st.session_state.remuneration_rates['mock_test']},
        {'Category': 'EY Personnel', 'Amount (₹)': st.session_state.remuneration_rates['ey_personnel']}
    ])
    
    st.dataframe(rates_df, use_container_width=True, hide_index=True)

def show_data_management():
    """Display data management options"""
    st.markdown("### 🛠️ Data Management")
    
    # Backup Management
    st.markdown("#### 💾 Backup Management")
    
    col_back1, col_back2 = st.columns(2)
    
    with col_back1:
        # Create backup
        backup_desc = st.text_input("Backup Description:", 
                                   placeholder="Optional description for backup")
        
        if st.button("🔒 Create New Backup", use_container_width=True):
            backup_file = create_backup(backup_desc)
            if backup_file:
                st.success(f"✅ Backup created: {backup_file.name}")
            else:
                st.error("❌ Failed to create backup")
    
    with col_back2:
        # List backups
        backup_files = list(BACKUP_DIR.glob("*.json"))
        if backup_files:
            backup_options = [f"{f.name} ({f.stat().st_size/1024:.1f} KB)" for f in sorted(backup_files, reverse=True)]
            selected_backup = st.selectbox("Select Backup:", backup_options)
            
            if st.button("🔄 Restore Backup", type="secondary", use_container_width=True):
                # Extract filename
                backup_filename = selected_backup.split(" (")[0]
                backup_file = BACKUP_DIR / backup_filename
                
                st.warning("⚠️ This will overwrite current data!")
                if st.checkbox("I understand this will overwrite current data"):
                    if restore_from_backup(backup_file):
                        st.success("✅ Backup restored successfully!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to restore backup")
        else:
            st.info("No backup files available")
    
    # Data Export
    st.markdown("---")
    st.markdown("#### 📤 Data Export")
    
    export_format = st.radio("Export Format:", ["CSV", "Excel", "JSON"])
    
    if st.button("📥 Export All System Data", use_container_width=True):
        # Prepare export data
        export_data = {
            'exam_data': st.session_state.exam_data,
            'allocation_references': st.session_state.allocation_references,
            'remuneration_rates': st.session_state.remuneration_rates,
            'deleted_records': st.session_state.deleted_records,
            'export_timestamp': datetime.now().isoformat(),
            'system_version': '2.0'
        }
        
        if export_format == "JSON":
            data_str = json.dumps(export_data, indent=4, default=str)
            mime_type = "application/json"
            extension = "json"
        else:
            # Create Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Exam data
                if st.session_state.exam_data:
                    exam_list = []
                    for exam_key, data in st.session_state.exam_data.items():
                        if isinstance(data, dict):
                            io_count = len(data.get('io_allocations', []))
                            ey_count = len(data.get('ey_allocations', []))
                        else:
                            io_count = len(data)
                            ey_count = 0
                        
                        exam_list.append({
                            'Exam': exam_key,
                            'IO Allocations': io_count,
                            'EY Allocations': ey_count
                        })
                    
                    if exam_list:
                        pd.DataFrame(exam_list).to_excel(writer, sheet_name='Exams', index=False)
                
                # Allocation references
                if st.session_state.allocation_references:
                    ref_list = []
                    for exam_key, roles in st.session_state.allocation_references.items():
                        for role, ref in roles.items():
                            ref_list.append({
                                'Exam': exam_key,
                                'Role': role,
                                'Order No.': ref.get('order_no', ''),
                                'Page No.': ref.get('page_no', '')
                            })
                    
                    if ref_list:
                        pd.DataFrame(ref_list).to_excel(writer, sheet_name='References', index=False)
                
                # Rates
                rates_df = pd.DataFrame([
                    {'Category': 'Multiple Shifts', 'Amount (₹)': st.session_state.remuneration_rates['multiple_shifts']},
                    {'Category': 'Single Shift', 'Amount (₹)': st.session_state.remuneration_rates['single_shift']},
                    {'Category': 'Mock Test', 'Amount (₹)': st.session_state.remuneration_rates['mock_test']},
                    {'Category': 'EY Personnel', 'Amount (₹)': st.session_state.remuneration_rates['ey_personnel']}
                ])
                rates_df.to_excel(writer, sheet_name='Rates', index=False)
            
            output.seek(0)
            data_str = output.getvalue()
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"
        
        # Download button
        filename = f"ssc_system_export_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"
        
        if export_format == "Excel":
            st.download_button(
                label="⬇️ Download Excel Export",
                data=data_str,
                file_name=filename,
                mime=mime_type,
                use_container_width=True
            )
        else:
            st.download_button(
                label=f"⬇️ Download {export_format} Export",
                data=data_str,
                file_name=filename,
                mime=mime_type,
                use_container_width=True
            )
    
    # Data Cleanup
    st.markdown("---")
    st.markdown("#### 🧹 Data Cleanup")
    
    if st.button("🗑️ Clear All Data", type="secondary"):
        st.error("⚠️ DANGER ZONE - This will delete ALL data!")
        
        col_warn1, col_warn2 = st.columns(2)
        with col_warn1:
            confirm1 = st.checkbox("I understand this will delete ALL exams")
        with col_warn2:
            confirm2 = st.checkbox("I understand this will delete ALL allocations")
        
        if confirm1 and confirm2:
            if st.button("🔥 CONFIRM DELETE ALL DATA", type="primary"):
                # Create final backup
                final_backup = create_backup("final_backup_before_wipe")
                
                # Clear all data
                st.session_state.exam_data = {}
                st.session_state.allocation = []
                st.session_state.ey_allocation = []
                st.session_state.allocation_references = {}
                st.session_state.deleted_records = []
                st.session_state.current_exam_key = ""
                
                # Clear files
                for file in [DATA_FILE, REFERENCE_FILE, DELETED_RECORDS_FILE]:
                    if file.exists():
                        file.unlink()
                
                # Clear backups (optional)
                clear_backups = st.checkbox("Also delete all backup files")
                if clear_backups:
                    for backup_file in BACKUP_DIR.glob("*.json"):
                        backup_file.unlink()
                
                save_all_data()
                
                if final_backup:
                    st.warning(f"✅ All data cleared. Final backup created: {final_backup.name}")
                else:
                    st.warning("✅ All data cleared. No backup created.")
                
                st.rerun()

def show_system_info():
    """Display system information"""
    st.markdown("### ℹ️ System Information")
    
    # System details
    info_data = {
        'System Name': 'SSC (ER) Kolkata - Allocation System',
        'Version': '2.0 (Streamlit Web Edition)',
        'Developer': 'Bijay Paswan',
        'Last Updated': datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
        'Python Version': sys.version.split()[0],
        'Streamlit Version': st.__version__,
        'Pandas Version': pd.__version__
    }
    
    for key, value in info_data.items():
        st.write(f"**{key}:** {value}")
    
    # System statistics
    st.markdown("---")
    st.markdown("#### 📊 System Statistics")
    
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    
    with col_stat1:
        st.metric("Total Exams", len(st.session_state.exam_data))
    
    with col_stat2:
        st.metric("Total Allocations", 
                 len(st.session_state.allocation) + len(st.session_state.ey_allocation))
    
    with col_stat3:
        # Calculate storage usage
        total_size = 0
        for file in DATA_DIR.glob("**/*"):
            if file.is_file():
                total_size += file.stat().st_size
        
        st.metric("Storage Used", f"{total_size/1024:.1f} KB")
    
    # Data files
    st.markdown("---")
    st.markdown("#### 📁 Data Files")
    
    data_files = []
    for file in DATA_DIR.glob("*"):
        if file.is_file():
            size_kb = file.stat().st_size / 1024
            modified = datetime.fromtimestamp(file.stat().st_mtime).strftime('%d-%m-%Y %H:%M')
            data_files.append({
                'File': file.name,
                'Size (KB)': f"{size_kb:.1f}",
                'Last Modified': modified
            })
    
    if data_files:
        st.table(data_files)
    else:
        st.info("No data files found")
    
    # Backup files
    st.markdown("#### 💾 Backup Files")
    
    backup_files = list(BACKUP_DIR.glob("*.json"))
    if backup_files:
        backup_info = []
        for file in sorted(backup_files, reverse=True)[:10]:  # Show last 10
            size_kb = file.stat().st_size / 1024
            modified = datetime.fromtimestamp(file.stat().st_mtime).strftime('%d-%m-%Y %H:%M')
            backup_info.append({
                'Backup': file.name,
                'Size (KB)': f"{size_kb:.1f}",
                'Created': modified
            })
        
        st.table(backup_info)
    else:
        st.info("No backup files found")

def show_help_support():
    """Display help and support information"""
    st.markdown("### 🆘 Help & Support")
    
    # Quick Start Guide
    with st.expander("🚀 Quick Start Guide", expanded=True):
        st.markdown("""
        **Step-by-Step Guide:**
        
        1. **Create Exam:**
           - Go to "Exam Management"
           - Enter exam name and year
           - Click "Create/Update Exam"
        
        2. **Load Master Data:**
           - Go to "Centre Coordinator" section
           - Load IO Master and Venue List files
           - Use default data for testing
        
        3. **Allocate Personnel:**
           - Select venue and dates
           - Choose Centre Coordinator or EY Personnel
           - Enter allocation reference
           - Click allocate
        
        4. **Generate Reports:**
           - Go to "Reports" section
           - View allocations and remuneration
           - Export data as needed
        """)
    
    # FAQ
    with st.expander("❓ Frequently Asked Questions"):
        st.markdown("""
        **Q: How do I import my existing data?**
        A: Use the file uploaders in each section to import Excel files with your data.
        
        **Q: Can I export data for offline use?**
        A: Yes! All sections have export options in CSV, Excel, and JSON formats.
        
        **Q: How are remuneration amounts calculated?**
        A: System uses the rates from Settings. Multiple shifts = higher rate, mock tests = separate rate.
        
        **Q: Is my data secure?**
        A: All data is stored locally in the `data/` directory. Regular backups are recommended.
        
        **Q: Can I use this on multiple computers?**
        A: Yes, export your data and import it on another computer.
        """)
    
    # Contact Information
    with st.expander("📞 Contact & Support"):
        st.markdown("""
        **For Technical Support:**
        - **Developer:** Bijay Paswan
        - **System:** SSC (ER) Kolkata Allocation System
        
        **Important Notes:**
        - This is a web application built with Streamlit
        - All data is stored locally
        - Regular backups are recommended
        - For feature requests or bug reports, contact the developer
        
        **Data Formats:**
        - IO Master: Requires NAME, AREA, CENTRE_CODE columns
        - Venue List: Requires VENUE, DATE, SHIFT columns
        - EY Master: Requires NAME column, other columns optional
        """)
    
    # System Status
    st.markdown("---")
    st.markdown("#### 🟢 System Status")
    
    # Check system health
    system_checks = []
    
    # Check data directory
    if DATA_DIR.exists():
        system_checks.append(("✅ Data Directory", "Accessible"))
    else:
        system_checks.append(("❌ Data Directory", "Not Found"))
    
    # Check write permissions
    try:
        test_file = DATA_DIR / "test.txt"
        test_file.write_text("test")
        test_file.unlink()
        system_checks.append(("✅ Write Permissions", "OK"))
    except:
        system_checks.append(("❌ Write Permissions", "Failed"))
    
    # Check data files
    for file, name in [(DATA_FILE, "Exam Data"), (REFERENCE_FILE, "References")]:
        if file.exists():
            size_kb = file.stat().st_size / 1024
            system_checks.append((f"✅ {name}", f"{size_kb:.1f} KB"))
        else:
            system_checks.append((f"⚠️ {name}", "Not Found"))
    
    # Display checks
    for check, status in system_checks:
        st.write(f"{check}: {status}")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def show_current_data_preview():
    """Display preview of current data"""
    st.markdown("### 📊 Current Data Preview")
    
    # IO Data
    if not st.session_state.io_df.empty:
        st.markdown("#### 👨‍💼 Centre Coordinator Data")
        st.dataframe(st.session_state.io_df.head(), use_container_width=True)
        st.write(f"**Total Records:** {len(st.session_state.io_df)}")
    else:
        st.info("No IO data loaded")
    
    # Venue Data
    if not st.session_state.venue_df.empty:
        st.markdown("#### 🏢 Venue Data")
        st.dataframe(st.session_state.venue_df.head(), use_container_width=True)
        st.write(f"**Total Records:** {len(st.session_state.venue_df)}")
    else:
        st.info("No venue data loaded")
    
    # EY Data
    if not st.session_state.ey_df.empty:
        st.markdown("#### 👁️ EY Personnel Data")
        st.dataframe(st.session_state.ey_df.head(), use_container_width=True)
        st.write(f"**Total Records:** {len(st.session_state.ey_df)}")
    else:
        st.info("No EY data loaded")

def show_allocation_references():
    """Display allocation references interface"""
    st.markdown("### 📚 Allocation References")
    
    if not st.session_state.allocation_references:
        st.info("No allocation references available")
        return
    
    # Create expandable sections for each exam
    for exam_key, roles in st.session_state.allocation_references.items():
        with st.expander(f"📖 {exam_key}", expanded=False):
            for role, ref in roles.items():
                col_ref1, col_ref2 = st.columns([3, 1])
                
                with col_ref1:
                    st.write(f"**{role}:**")
                    st.write(f"  Order No.: {ref.get('order_no', 'N/A')}")
                    st.write(f"  Page No.: {ref.get('page_no', 'N/A')}")
                    
                    remarks = ref.get('remarks', '')
                    if remarks:
                        st.write(f"  Remarks: {remarks}")
                    
                    timestamp = ref.get('timestamp', '')
                    if timestamp:
                        try:
                            timestamp = datetime.fromisoformat(timestamp).strftime("%d-%m-%Y %H:%M")
                            st.write(f"  Created: {timestamp}")
                        except:
                            pass
                
                with col_ref2:
                    if st.button("🗑️ Delete", key=f"del_ref_{exam_key}_{role}"):
                        if st.checkbox(f"Confirm delete reference for {role}"):
                            del st.session_state.allocation_references[exam_key][role]
                            
                            # Remove exam if no references left
                            if not st.session_state.allocation_references[exam_key]:
                                del st.session_state.allocation_references[exam_key]
                            
                            save_all_data()
                            st.success("Reference deleted!")
                            st.rerun()

def generate_io_report():
    """Generate comprehensive IO report"""
    if not st.session_state.allocation:
        st.warning("No allocation data available")
        return
    
    alloc_df = pd.DataFrame(st.session_state.allocation)
    
    # Create detailed report
    report_data = []
    
    for io_name in alloc_df['IO Name'].unique():
        io_data = alloc_df[alloc_df['IO Name'] == io_name]
        
        # Calculate statistics
        total_days = io_data['Date'].nunique()
        total_shifts = len(io_data)
        total_venues = io_data['Venue'].nunique()
        
        # Count by shift type
        mock_days = io_data[io_data['Mock Test']]['Date'].nunique()
        regular_days = total_days - mock_days
        
        # Group by date to find multiple shifts
        date_shift_counts = io_data.groupby('Date')['Shift'].nunique()
        multiple_shift_days = (date_shift_counts > 1).sum()
        single_shift_days = (date_shift_counts == 1).sum()
        
        # Calculate remuneration
        mock_amount = mock_days * st.session_state.remuneration_rates['mock_test']
        single_amount = single_shift_days * st.session_state.remuneration_rates['single_shift']
        multi_amount = multiple_shift_days * st.session_state.remuneration_rates['multiple_shifts']
        total_amount = mock_amount + single_amount + multi_amount
        
        report_data.append({
            'IO Name': io_name,
            'Total Days': total_days,
            'Total Shifts': total_shifts,
            'Total Venues': total_venues,
            'Mock Days': mock_days,
            'Regular Days': regular_days,
            'Single Shift Days': single_shift_days,
            'Multiple Shift Days': multiple_shift_days,
            'Total Amount (₹)': total_amount
        })
    
    if report_data:
        report_df = pd.DataFrame(report_data)
        
        st.markdown("### 📊 IO Allocation Report")
        st.dataframe(report_df, use_container_width=True)
        
        # Export
        csv = report_df.to_csv(index=False)
        st.download_button(
            label="⬇️ Download Report",
            data=csv,
            file_name=f"io_allocation_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True
        )

# ============================================================================
# MAIN APPLICATION
# ============================================================================

def main():
    """Main application entry point"""
    try:
        # Configure page
        st.set_page_config(
            page_title="SSC (ER) Kolkata - Allocation System",
            page_icon="🏛️",
            layout="wide",
            initial_sidebar_state="expanded",
            menu_items={
                'Get Help': 'https://www.example.com',
                'Report a bug': 'https://www.example.com',
                'About': "### SSC (ER) Kolkata Allocation System\n\nVersion 2.0\n\nDesigned by Bijay Paswan"
            }
        )
        
        # Apply custom CSS
        st.markdown("""
            <style>
            /* Main styling */
            .main-header {
                text-align: center;
                padding: 1.5rem 0;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border-radius: 10px;
                margin-bottom: 2rem;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            }
            
            /* Card styling */
            .card {
                background: white;
                padding: 1.5rem;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                margin-bottom: 1rem;
                transition: transform 0.3s ease;
            }
            
            .card:hover {
                transform: translateY(-2px);
                box-shadow: 0 4px 20px rgba(0,0,0,0.15);
            }
            
            /* Metric cards */
            .metric-card {
                background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
                padding: 1.2rem;
                border-radius: 10px;
                text-align: center;
                box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                border: 1px solid #e0e0e0;
            }
            
            /* Button styling */
            .stButton > button {
                border-radius: 8px;
                font-weight: 500;
                transition: all 0.3s ease;
            }
            
            .stButton > button:hover {
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            }
            
            /* Success button */
            .success-button {
                background: linear-gradient(135deg, #56ab2f 0%, #a8e063 100%);
                color: white;
                border: none;
            }
            
            /* Danger button */
            .danger-button {
                background: linear-gradient(135deg, #ff416c 0%, #ff4b2b 100%);
                color: white;
                border: none;
            }
            
            /* Warning button */
            .warning-button {
                background: linear-gradient(135deg, #f46b45 0%, #eea849 100%);
                color: white;
                border: none;
            }
            
            /* Info button */
            .info-button {
                background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
                color: white;
                border: none;
            }
            
            /* Tab styling */
            .stTabs [data-baseweb="tab-list"] {
                gap: 8px;
            }
            
            .stTabs [data-baseweb="tab"] {
                height: 50px;
                white-space: pre-wrap;
                background-color: #f8f9fa;
                border-radius: 8px 8px 0 0;
                font-weight: 500;
            }
            
            .stTabs [aria-selected="true"] {
                background-color: #4169e1;
                color: white;
            }
            
            /* Dataframe styling */
            .dataframe {
                border-radius: 8px;
                overflow: hidden;
            }
            
            /* Sidebar styling */
            [data-testid="stSidebar"] {
                background: linear-gradient(180deg, #2c3e50 0%, #34495e 100%);
            }
            
            [data-testid="stSidebar"] * {
                color: white !important;
            }
            
            /* Hide Streamlit branding */
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            
            /* Scrollbar styling */
            ::-webkit-scrollbar {
                width: 8px;
                height: 8px;
            }
            
            ::-webkit-scrollbar-track {
                background: #f1f1f1;
                border-radius: 4px;
            }
            
            ::-webkit-scrollbar-thumb {
                background: #888;
                border-radius: 4px;
            }
            
            ::-webkit-scrollbar-thumb:hover {
                background: #555;
            }
            
            /* Responsive design */
            @media (max-width: 768px) {
                .main-header {
                    padding: 1rem;
                    font-size: 0.9rem;
                }
                
                .card {
                    padding: 1rem;
                }
            }
            </style>
        """, unsafe_allow_html=True)
        
        # Initialize session state (COMPLETELY)
        initialize_session_state()
        
        # Load existing data
        load_all_data()
        
        # Sidebar navigation
        with st.sidebar:
            st.markdown("""
                <div style='text-align: center; padding: 20px 0;'>
                    <h2>📋 Navigation</h2>
                    <p style='font-size: 0.9rem; color: #bdc3c7;'>SSC Allocation System</p>
                </div>
            """, unsafe_allow_html=True)
            
            # Menu selection
            menu_options = {
                "🏠 Dashboard": "dashboard",
                "📝 Exam Management": "exam",
                "👨‍💼 Centre Coordinator": "io",
                "👁️ EY Personnel": "ey",
                "📊 Reports": "reports",
                "⚙️ Settings": "settings"
            }
            
            selected_menu = st.radio(
                "Select Module:",
                list(menu_options.keys()),
                label_visibility="collapsed"
            )
            
            # Update session state
            st.session_state.menu = menu_options[selected_menu]
            
            st.markdown("---")
            
            # Current exam info
            st.markdown("### 🎯 Current Exam")
            if st.session_state.current_exam_key:
                st.success(f"**{st.session_state.current_exam_key[:30]}{'...' if len(st.session_state.current_exam_key) > 30 else ''}**")
                
                # Quick stats
                current_io = len([a for a in st.session_state.allocation 
                                if a.get('Exam') == st.session_state.current_exam_key])
                current_ey = len([a for a in st.session_state.ey_allocation 
                                if a.get('Exam') == st.session_state.current_exam_key])
                
                col_stat1, col_stat2 = st.columns(2)
                with col_stat1:
                    st.metric("IO", current_io)
                with col_stat2:
                    st.metric("EY", current_ey)
            else:
                st.warning("No exam selected")
            
            st.markdown("---")
            
            # Quick actions
            st.markdown("### ⚡ Quick Actions")
            
            col_q1, col_q2 = st.columns(2)
            with col_q1:
                if st.button("💾 Save", use_container_width=True):
                    save_all_data()
                    st.success("Data saved!")
            
            with col_q2:
                if st.button("🔄 Refresh", use_container_width=True):
                    st.rerun()
            
            if st.button("📥 Load Defaults", use_container_width=True):
                load_default_master_data()
                st.rerun()
            
            st.markdown("---")
            
            # System info
            st.markdown("### ℹ️ System Info")
            st.info(f"Updated: {datetime.now().strftime('%H:%M')}")
            
            # Developer credit
            st.markdown("---")
            st.markdown("""
                <div style='text-align: center; padding: 10px 0;'>
                    <p style='font-size: 0.8rem; color: #95a5a6;'>
                        Designed by Bijay Paswan<br>
                        SSC (ER) Kolkata
                    </p>
                </div>
            """, unsafe_allow_html=True)
        
        # Main content area
        try:
            # Display selected module
            if st.session_state.menu == "dashboard":
                show_dashboard()
            elif st.session_state.menu == "exam":
                show_exam_management()
            elif st.session_state.menu == "io":
                show_centre_coordinator()
            elif st.session_state.menu == "ey":
                show_ey_personnel()
            elif st.session_state.menu == "reports":
                show_reports()
            elif st.session_state.menu == "settings":
                show_settings()
            else:
                show_dashboard()
                
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            logging.error(f"Application error: {str(e)}")
            logging.error(traceback.format_exc())
            
            # Show error details in expander
            with st.expander("Error Details"):
                st.code(traceback.format_exc())
            
            # Recovery option
            if st.button("🔄 Restart Application"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
    
    except Exception as e:
        # Critical error handling
        st.error(f"Critical error: {str(e)}")
        logging.critical(f"Critical application error: {str(e)}")
        logging.critical(traceback.format_exc())

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    main()
